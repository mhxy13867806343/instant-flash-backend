from __future__ import annotations

import hashlib
import io
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path as FilePath
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Path, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel, Field
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from app.api.serializers import media_type
from app.api.user_identity import normalize_client_subtype, normalize_client_type, normalize_phone, phone_from_user_id
from app.api.utils import new_business_id
from app.core.security import create_access_token, decode_access_token, revoke_access_token
from app.db.base import utc_now
from app.db.session import get_db
from app.models.admin_agreement import AdminAgreement
from app.models.comment import Comment
from app.models.message import Message
from app.models.post import Post
from app.models.post_like import PostLike
from app.models.post_share import PostShare
from app.models.system_config import AdminAccessRule, AdminAccount, AdminAnnouncement, AdminDictionary, AdminFeedbackField, AdminFeedbackFormConfig, AdminMenu, AdminOperationLog, AdminPackage, AdminPermission, AdminRegion, AdminRole, AdminSecuritySetting, AdminSystemMessage, AdminTag, AdminVersion, FeedbackSubmission
from app.models.user import User

router = APIRouter(prefix="/api/admin", tags=["后台管理"])
admin_bearer = HTTPBearer(auto_error=False)


DEFAULT_AGREEMENTS = {
    "privacy": "<h2>即闪隐私政策</h2><p>请在后台编辑最新隐私政策内容。</p>",
    "user": "<h2>即闪用户协议</h2><p>请在后台编辑最新用户协议内容。</p>",
}
SINGLE_BANNER_ANNOUNCEMENT_ID = "SINGLE_BANNER"
PACKAGE_UPLOAD_ROOT = FilePath(__file__).resolve().parents[2] / "static" / "uploads" / "packages"
USER_ONLINE_WINDOW_SECONDS = 5 * 60
DEFAULT_FEEDBACK_CONFIG_ID = "default"
DEFAULT_FEEDBACK_FIELDS = [
    {"field_id": "fb_field_phone", "field_key": "phone", "label": "手机号码", "type": "phone", "placeholder": "请填写手机号码", "required": True, "sort": 10},
    {"field_id": "fb_field_title", "field_key": "title", "label": "反馈标题", "type": "input", "placeholder": "请输入标题或菜单名称", "required": True, "sort": 20},
    {"field_id": "fb_field_content", "field_key": "content", "label": "反馈内容", "type": "textarea", "placeholder": "请填写具体反馈内容", "required": True, "sort": 30},
]
DEFAULT_ADMIN_PERMISSIONS = ["dashboard", "user", "content", "comment", "like", "share", "simulator", "account", "announcement", "version", "system", "tag", "region", "dict", "menu", "message", "agreement", "log", "access_rule", "user_online", "feedback"]
DEFAULT_ADMIN_PERMISSION_MODULES: list[dict[str, Any]] = [
    {"permission_id": "perm_dashboard", "permission_key": "dashboard", "label": "数据看板", "description": "后台首页数据看板", "sort": 10, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_user", "permission_key": "user", "label": "用户管理", "description": "用户列表、禁用和详情", "sort": 20, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_content", "permission_key": "content", "label": "内容管理", "description": "动态内容审核与上下架", "sort": 30, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_comment", "permission_key": "comment", "label": "评论管理", "description": "评论列表与删除", "sort": 40, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_like", "permission_key": "like", "label": "点赞管理", "description": "内容点赞记录管理", "sort": 45, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_share", "permission_key": "share", "label": "分享管理", "description": "内容分享记录管理", "sort": 46, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_simulator", "permission_key": "simulator", "label": "App 仿真模拟", "description": "App 仿真模拟页面", "sort": 50, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_account", "permission_key": "account", "label": "账号管理", "description": "后台账号和角色配置", "sort": 60, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_announcement", "permission_key": "announcement", "label": "公告管理", "description": "单公告和公告列表", "sort": 70, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_version", "permission_key": "version", "label": "版本管理", "description": "App 版本管理", "sort": 80, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_system", "permission_key": "system", "label": "系统配置", "description": "系统配置目录权限", "sort": 90, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_tag", "permission_key": "tag", "label": "标签管理", "description": "动态标签配置", "sort": 100, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_region", "permission_key": "region", "label": "地区管理", "description": "地区和定位配置", "sort": 110, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_dict", "permission_key": "dict", "label": "字典管理", "description": "业务字典配置", "sort": 120, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_menu", "permission_key": "menu", "label": "菜单管理", "description": "后台动态菜单配置", "sort": 130, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_message", "permission_key": "message", "label": "系统消息", "description": "系统消息推送", "sort": 140, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_agreement", "permission_key": "agreement", "label": "协议管理", "description": "用户协议和隐私协议", "sort": 150, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_log", "permission_key": "log", "label": "日志管理", "description": "后台登录日志与操作日志", "sort": 160, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_access_rule", "permission_key": "access_rule", "label": "黑白名单", "description": "接口限流黑名单与白名单配置", "sort": 170, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_user_online", "permission_key": "user_online", "label": "在线用户", "description": "查看用户端在线和离线用户", "sort": 180, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
    {"permission_id": "perm_feedback", "permission_key": "feedback", "label": "反馈管理", "description": "配置用户端反馈动态表单并处理反馈", "sort": 190, "status": "enabled", "is_default": True, "remark": "系统默认权限"},
]
DEFAULT_ADMIN_ROLES: list[dict[str, Any]] = [
    {"role_id": "role_superadmin", "role_key": "superadmin", "label": "超级管理员", "icon": "StarFilled", "permissions": DEFAULT_ADMIN_PERMISSIONS, "sort": 10, "status": "enabled", "is_default": True, "remark": "系统内置超级管理员角色"},
    {"role_id": "role_admin", "role_key": "admin", "label": "管理员", "icon": "UserFilled", "permissions": ["dashboard", "user", "content", "comment", "like", "share", "account", "tag", "region", "message", "log", "access_rule", "user_online", "feedback"], "sort": 20, "status": "enabled", "is_default": True, "remark": "系统内置管理员角色"},
    {"role_id": "role_operator", "role_key": "operator", "label": "运营员", "icon": "Setting", "permissions": ["dashboard", "content", "comment", "like", "share", "tag"], "sort": 30, "status": "enabled", "is_default": True, "remark": "系统内置运营角色"},
    {"role_id": "role_viewer", "role_key": "viewer", "label": "观察员", "icon": "View", "permissions": ["dashboard"], "sort": 40, "status": "enabled", "is_default": True, "remark": "系统内置观察员角色"},
]
DEFAULT_ADMIN_MENUS: list[dict[str, Any]] = [
    {"menu_id": "menu_dashboard", "parent_id": None, "title": "数据看板", "path": "/dashboard", "name": "Dashboard", "component": "views/dashboard/Index", "icon": "Odometer", "type": "menu", "permission": "dashboard", "sort": 10, "affix": True, "remark": "后台首页数据看板"},
    {"menu_id": "menu_user", "parent_id": None, "title": "用户管理", "path": "/user", "name": "UserList", "component": "views/user/List", "icon": "User", "type": "menu", "permission": "user", "sort": 20},
    {"menu_id": "menu_content", "parent_id": None, "title": "内容管理", "path": "/content", "name": "ContentList", "component": "views/content/List", "icon": "Document", "type": "menu", "permission": "content", "sort": 30},
    {"menu_id": "menu_comment", "parent_id": None, "title": "评论管理", "path": "/comment", "name": "CommentList", "component": "views/comment/List", "icon": "ChatLineSquare", "type": "menu", "permission": "comment", "sort": 40},
    {"menu_id": "menu_like", "parent_id": None, "title": "点赞管理", "path": "/like", "name": "LikeList", "component": "views/like/List", "icon": "Star", "type": "menu", "permission": "like", "sort": 45},
    {"menu_id": "menu_share", "parent_id": None, "title": "分享管理", "path": "/share", "name": "ShareList", "component": "views/share/List", "icon": "Share", "type": "menu", "permission": "share", "sort": 46},
    {"menu_id": "menu_simulator", "parent_id": None, "title": "App 仿真模拟", "path": "/simulator", "name": "AppSimulator", "component": "views/simulator/Index", "icon": "Smartphone", "type": "menu", "permission": "simulator", "sort": 50},
    {"menu_id": "menu_account", "parent_id": None, "title": "账号管理", "path": "/account", "name": "AccountList", "component": "views/account/List", "icon": "UserFilled", "type": "menu", "permission": "account", "sort": 60},
    {"menu_id": "menu_announcement", "parent_id": None, "title": "公告管理", "path": "/announcement", "name": "Announcement", "component": None, "redirect": "/announcement/single", "icon": "Bell", "type": "catalog", "permission": "announcement", "sort": 70},
    {"menu_id": "menu_announcement_single", "parent_id": "menu_announcement", "title": "单公告", "path": "/announcement/single", "name": "AnnouncementSingle", "component": "views/announcement/Single", "icon": "Promotion", "type": "menu", "permission": "announcement", "sort": 10},
    {"menu_id": "menu_announcement_list", "parent_id": "menu_announcement", "title": "公告列表", "path": "/announcement/list", "name": "AnnouncementList", "component": "views/announcement/List", "icon": "List", "type": "menu", "permission": "announcement", "sort": 20},
    {"menu_id": "menu_version", "parent_id": None, "title": "版本管理", "path": "/version", "name": "VersionList", "component": "views/version/List", "icon": "Upload", "type": "menu", "permission": "version", "sort": 80},
    {"menu_id": "menu_version_packages", "parent_id": None, "title": "安装包管理", "path": "/version/packages", "name": "VersionPackageUpload", "component": "views/version/Packages", "icon": "UploadFilled", "type": "menu", "permission": "version", "sort": 81},
    {"menu_id": "menu_account_profile", "parent_id": None, "title": "个人信息", "path": "account/profile", "name": "AccountProfile", "component": "views/account/Profile", "icon": "User", "type": "menu", "permission": None, "sort": 85, "visible": False, "remark": "个人中心动态路由，仅管理员和超级管理员可见"},
    {"menu_id": "menu_account_settings", "parent_id": None, "title": "安全设置", "path": "account/settings", "name": "AccountSettings", "component": "views/account/Settings", "icon": "Lock", "type": "menu", "permission": None, "sort": 86, "visible": False, "remark": "个人中心动态路由，仅管理员和超级管理员可见"},
    {"menu_id": "menu_system", "parent_id": None, "title": "系统配置", "path": "/system", "name": "SystemConfig", "component": None, "redirect": "/tag", "icon": "Setting", "type": "catalog", "permission": None, "sort": 90},
    {"menu_id": "menu_tag", "parent_id": "menu_system", "title": "标签管理", "path": "/tag", "name": "TagList", "component": "views/tag/List", "icon": "PriceTag", "type": "menu", "permission": "tag", "sort": 10},
    {"menu_id": "menu_region", "parent_id": "menu_system", "title": "地区管理", "path": "/region", "name": "RegionList", "component": "views/region/List", "icon": "Location", "type": "menu", "permission": "region", "sort": 20},
    {"menu_id": "menu_dict", "parent_id": "menu_system", "title": "字典管理", "path": "/dict", "name": "DictList", "component": "views/dict/List", "icon": "Memo", "type": "menu", "permission": "dict", "sort": 30},
    {"menu_id": "menu_menu", "parent_id": "menu_system", "title": "菜单管理", "path": "/menu", "name": "MenuList", "component": "views/menu/List", "icon": "Menu", "type": "menu", "permission": "menu", "sort": 40},
    {"menu_id": "menu_message", "parent_id": "menu_system", "title": "系统消息", "path": "/message", "name": "SysMessage", "component": "views/message/List", "icon": "Message", "type": "menu", "permission": "message", "sort": 50},
    {"menu_id": "menu_privacy", "parent_id": "menu_system", "title": "隐私协议", "path": "/agreement/privacy", "name": "PrivacyAgreement", "component": "views/agreement/Privacy", "icon": "Lock", "type": "menu", "permission": "agreement", "sort": 60},
    {"menu_id": "menu_user_agreement", "parent_id": "menu_system", "title": "用户协议", "path": "/agreement/user", "name": "UserAgreement", "component": "views/agreement/User", "icon": "Checked", "type": "menu", "permission": "agreement", "sort": 70},
    {"menu_id": "menu_log", "parent_id": "menu_system", "title": "日志管理", "path": "log/list", "name": "LogList", "component": "views/log/List", "icon": "Tickets", "type": "menu", "permission": "log", "sort": 80},
    {"menu_id": "menu_access_rule", "parent_id": "menu_system", "title": "黑白名单", "path": "security/access-rules", "name": "AccessRuleList", "component": "views/security/AccessRules", "icon": "Connection", "type": "menu", "permission": "access_rule", "sort": 90},
    {"menu_id": "menu_user_online", "parent_id": "menu_system", "title": "在线用户", "path": "user/online", "name": "UserOnline", "component": "views/user/Online", "icon": "Monitor", "type": "menu", "permission": "user_online", "sort": 100},
    {"menu_id": "menu_feedback", "parent_id": "menu_system", "title": "反馈管理", "path": "feedback", "name": "FeedbackManage", "component": "views/feedback/List", "icon": "ChatDotRound", "type": "menu", "permission": "feedback", "sort": 110},
]
ADMIN_ONLY_ROUTE_NAMES = {"AccountProfile", "AccountSettings", "LogList"}
ADMIN_ROUTE_ROLES = {"admin", "superadmin"}


class AdminLoginRequest(BaseModel):
    username: str = Field(min_length=1, max_length=64, title="管理员账号", description="后台管理员登录账号")
    password: str = Field(min_length=1, max_length=128, title="管理员密码", description="后台管理员登录密码")


class AdminUserStatusUpdate(BaseModel):
    status: str = Field(pattern="^(normal|banned)$", title="用户状态", description="normal 表示正常，banned 表示禁用")


class AdminSecuritySettingPayload(BaseModel):
    mfaEnabled: bool | None = Field(default=None, title="二次登录验证", description="是否开启二次登录验证")
    passwordPolicyEnabled: bool | None = Field(default=None, title="密码策略审核", description="是否开启密码策略审核")
    remark: str | None = Field(default=None, title="备注", description="安全设置备注")


class AdminAccessRulePayload(BaseModel):
    type: str = Field(pattern="^(blacklist|whitelist)$", title="名单类型", description="blacklist 黑名单，whitelist 白名单")
    ip: str | None = Field(default=None, max_length=64, title="IP 地址", description="IP 地址，支持 * 通配；为空表示全部 IP")
    method: str | None = Field(default=None, pattern="^(ALL|GET|POST|PUT|PATCH|DELETE|OPTIONS|HEAD)$", title="请求方法", description="ALL 表示全部方法")
    path: str | None = Field(default=None, max_length=256, title="接口路径", description="接口路径，支持 /api/posts/* 这种前缀通配；为空表示全部接口")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")
    remark: str | None = Field(default=None, title="备注", description="规则备注")


class AgreementUpdate(BaseModel):
    content: str = Field(title="协议内容", description="HTML 格式的协议正文")


class AdminTagPayload(BaseModel):
    name: str = Field(min_length=1, max_length=64, title="标签名称", description="标签管理中的标签名称")
    color: str | None = Field(default=None, max_length=32, title="标签颜色", description="标签颜色，例如 #1677ff")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")
    remark: str | None = Field(default=None, title="备注", description="标签备注")


class AdminRegionPayload(BaseModel):
    name: str = Field(min_length=1, max_length=64, title="地区名称", description="省市区名称")
    code: str = Field(min_length=1, max_length=32, title="地区编码", description="行政区划编码或前端自定义编码")
    parentId: str | None = Field(default=None, max_length=64, title="上级地区 ID", description="上级业务地区 ID，顶级地区可为空")
    level: int = Field(default=1, ge=1, le=4, title="层级", description="1 省级，2 市级，3 区县，4 街道")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")


class AdminDictionaryPayload(BaseModel):
    type: str = Field(min_length=1, max_length=64, title="字典类型", description="字典分组编码，例如 post_status")
    parentId: str | None = Field(default=None, max_length=64, title="上级字典 ID", description="上级字典项的业务 ID；为空表示顶级字典项")
    label: str = Field(min_length=1, max_length=128, title="字典标签", description="展示给用户看的中文名称")
    value: str = Field(min_length=1, max_length=128, title="字典值", description="前后端传递使用的值")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")
    remark: str | None = Field(default=None, title="备注", description="字典项备注")


class AdminAnnouncementPayload(BaseModel):
    title: str | None = Field(default=None, max_length=128, title="公告标题", description="公告标题")
    type: str | None = Field(default=None, pattern="^(info|warning|danger)$", title="公告类型", description="info 普通，warning 重要提醒，danger 紧急公告")
    content: str | None = Field(default=None, title="公告内容", description="公告正文，支持富文本")
    link: str | None = Field(default=None, max_length=512, title="跳转链接", description="公告关联链接")
    pinned: bool | None = Field(default=None, title="是否置顶", description="是否置顶展示")
    startTime: str | None = Field(default=None, max_length=32, title="开始时间", description="公告生效开始时间")
    endTime: str | None = Field(default=None, max_length=32, title="结束时间", description="公告生效结束时间")
    status: str | None = Field(default=None, pattern="^(active|inactive)$", title="状态", description="active 已发布，inactive 已停用")


class AdminVersionPayload(BaseModel):
    platform: str | None = Field(default=None, pattern="^(iOS|Android|HarmonyOS)$", title="平台", description="版本平台")
    version: str | None = Field(default=None, max_length=32, title="版本号", description="展示版本号")
    build: str | None = Field(default=None, max_length=32, title="Build 号", description="内部构建号")
    forceUpdate: bool | None = Field(default=None, title="是否强制更新", description="是否要求客户端强制更新")
    status: str | None = Field(default=None, pattern="^(released|beta|deprecated)$", title="状态", description="released 已发布，beta 灰度中，deprecated 已下线")
    betaPct: int | None = Field(default=None, ge=0, le=100, title="灰度比例", description="灰度发布比例")
    notes: str | None = Field(default=None, title="更新说明", description="更新说明，支持文本或富文本")
    notesType: str | None = Field(default=None, pattern="^(text|rich)$", title="说明类型", description="text 文本，rich 富文本")
    downloadUrl: str | None = Field(default=None, max_length=512, title="下载地址", description="安装包下载地址")


class AdminPackageUpdatePayload(BaseModel):
    displayName: str | None = Field(default=None, max_length=256, title="安装包展示名称", description="前端列表展示的包名称，可修改")
    version: str | None = Field(default=None, max_length=32, title="版本号", description="安装包版本号")
    build: str | None = Field(default=None, max_length=32, title="Build 号", description="安装包构建号")
    status: str | None = Field(default=None, pattern="^(uploaded|active|deprecated)$", title="状态", description="uploaded 已上传，active 启用，deprecated 已下线")
    remark: str | None = Field(default=None, title="备注", description="安装包备注")


class AdminAccountPayload(BaseModel):
    username: str | None = Field(default=None, max_length=64, title="登录账号", description="后台账号用户名")
    nickname: str | None = Field(default=None, max_length=64, title="昵称", description="后台账号昵称")
    avatar: str | None = Field(default=None, max_length=512, title="头像", description="头像地址")
    role: str | None = Field(default=None, max_length=64, title="角色", description="后台角色 key，来自角色管理")
    permissions: list[str] | None = Field(default=None, title="权限模块", description="账号可访问的权限模块")
    status: str | None = Field(default=None, pattern="^(active|disabled)$", title="状态", description="active 正常，disabled 禁用")
    email: str | None = Field(default=None, max_length=128, title="邮箱", description="管理员邮箱")
    phone: str | None = Field(default=None, max_length=32, title="手机号", description="管理员手机号")
    remark: str | None = Field(default=None, title="备注", description="账号备注")


class AdminRolePayload(BaseModel):
    roleKey: str = Field(min_length=1, max_length=64, title="角色标识", description="角色 key，例如 operator")
    label: str = Field(min_length=1, max_length=64, title="角色名称", description="角色中文名称")
    icon: str | None = Field(default=None, max_length=64, title="角色图标", description="前端角色图标名")
    permissions: list[str] = Field(default_factory=list, title="权限模块", description="角色默认权限模块")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")
    remark: str | None = Field(default=None, title="备注", description="角色备注")


class AdminPermissionPayload(BaseModel):
    permissionKey: str = Field(min_length=1, max_length=64, title="权限标识", description="权限模块 key，例如 dashboard")
    label: str = Field(min_length=1, max_length=64, title="权限名称", description="权限模块中文名称")
    description: str | None = Field(default=None, max_length=256, title="权限说明", description="权限模块用途说明")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")
    remark: str | None = Field(default=None, title="备注", description="权限模块备注")


class AdminMenuPayload(BaseModel):
    parentId: str | None = Field(default=None, max_length=64, title="上级菜单 ID", description="父级菜单业务 ID；顶级菜单为空")
    title: str = Field(min_length=1, max_length=64, title="菜单标题", description="菜单展示中文名称")
    path: str = Field(min_length=1, max_length=128, title="路由路径", description="前端路由 path，例如 /dashboard")
    name: str = Field(min_length=1, max_length=64, title="路由名称", description="前端路由 name，必须唯一")
    component: str | None = Field(default=None, max_length=128, title="组件路径", description="前端组件路径，例如 views/dashboard/Index")
    redirect: str | None = Field(default=None, max_length=128, title="重定向地址", description="父级菜单默认跳转地址")
    icon: str | None = Field(default=None, max_length=64, title="图标", description="Element Plus 图标组件名")
    type: str = Field(default="menu", pattern="^(catalog|menu|button|link)$", title="菜单类型", description="catalog 目录，menu 菜单，button 按钮，link 外链")
    permission: str | None = Field(default=None, max_length=64, title="权限标识", description="账号权限模块，例如 dashboard/user/content")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 禁用")
    visible: bool = Field(default=True, title="是否显示", description="是否在侧边菜单显示")
    keepAlive: bool = Field(default=False, title="是否缓存", description="前端 keepAlive 元信息")
    affix: bool = Field(default=False, title="是否固定页签", description="前端 affix 元信息")
    externalLink: str | None = Field(default=None, max_length=512, title="外链地址", description="外链菜单地址")
    remark: str | None = Field(default=None, title="备注", description="菜单备注")


class BatchIdsPayload(BaseModel):
    ids: list[str] = Field(default_factory=list, title="ID 列表", description="需要批量操作的业务 ID 列表")


class AdminSystemMessagePayload(BaseModel):
    title: str = Field(min_length=1, max_length=128, title="消息标题", description="系统消息标题")
    content: str = Field(min_length=1, title="消息内容", description="系统消息正文")
    type: str = Field(default="notice", max_length=32, title="消息类型", description="notice 通知，warning 警告，activity 活动")
    target: str = Field(default="all", max_length=32, title="发送范围", description="all 全部用户，admin 后台，user 用户端")
    status: str = Field(default="draft", pattern="^(draft|published|disabled)$", title="状态", description="draft 草稿，published 已发布，disabled 已停用")
    isPinned: bool = Field(default=False, title="是否置顶", description="是否在系统消息列表置顶展示")


class AdminFeedbackConfigPayload(BaseModel):
    title: str = Field(default="意见反馈", min_length=1, max_length=128, title="反馈页面标题", description="用户端反馈页面展示标题")
    menuTitle: str = Field(default="反馈管理", min_length=1, max_length=128, title="菜单标题", description="后台反馈管理菜单或页面标题")
    description: str | None = Field(default=None, title="页面说明", description="用户端反馈页面说明文案")
    submitButtonText: str = Field(default="提交反馈", min_length=1, max_length=64, title="提交按钮文案", description="用户端动态表单提交按钮文字")
    successMessage: str = Field(default="反馈提交成功", min_length=1, max_length=128, title="成功提示", description="用户提交成功后的提示文案")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 停用")
    remark: str | None = Field(default=None, title="备注", description="后台备注")


class AdminFeedbackFieldPayload(BaseModel):
    fieldKey: str = Field(min_length=1, max_length=64, title="字段标识", description="动态表单字段 key，例如 phone/title/content")
    label: str = Field(min_length=1, max_length=128, title="字段名称", description="表单展示名称")
    type: str = Field(default="input", pattern="^(input|textarea|select|radio|checkbox|number|phone|email|autocomplete|input_number|cascader|switch|rate|color|date|datetime|time)$", title="字段类型", description="动态表单字段类型")
    placeholder: str | None = Field(default=None, max_length=256, title="占位提示", description="输入框 placeholder")
    required: bool = Field(default=False, title="是否必填", description="是否为必填项")
    options: list[dict[str, Any]] = Field(default_factory=list, title="选项", description="select/radio/checkbox 使用的选项列表")
    sort: int = Field(default=0, ge=0, title="排序值", description="数字越小越靠前")
    status: str = Field(default="enabled", pattern="^(enabled|disabled)$", title="状态", description="enabled 启用，disabled 停用")
    remark: str | None = Field(default=None, title="备注", description="字段备注")


class AdminFeedbackStatusPayload(BaseModel):
    status: str = Field(pattern="^(pending|processing|resolved|rejected)$", title="处理状态", description="pending 待处理，processing 处理中，resolved 已解决，rejected 已驳回")
    reply: str | None = Field(default=None, title="处理回复", description="后台处理回复")
    remark: str | None = Field(default=None, title="备注", description="后台备注")


class AdminResponse(BaseModel):
    code: int = Field(title="业务状态码", description="200 表示成功，其他值表示业务失败")
    message: str = Field(title="提示信息", description="接口处理结果说明")
    data: Any = Field(default=None, title="响应数据", description="接口返回的业务数据")


def ok(data: Any = None, message: str = "success") -> dict[str, Any]:
    return {"code": 200, "message": message, "data": data}


def fail(status_code: int, message: str) -> HTTPException:
    return HTTPException(
        status_code=status_code,
        detail={"code": status_code, "message": message, "data": {}},
    )


def format_time(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def get_admin_subject(
    credentials: Annotated[HTTPAuthorizationCredentials | None, Depends(admin_bearer)],
) -> str:
    if credentials is None:
        raise fail(status.HTTP_401_UNAUTHORIZED, "未登录，请先登录")
    if credentials.scheme.lower() != "bearer":
        raise fail(status.HTTP_401_UNAUTHORIZED, "登录失效，请重新登录")
    subject = decode_access_token(credentials.credentials)
    if subject is None or not subject.startswith("admin:"):
        raise fail(status.HTTP_401_UNAUTHORIZED, "登录失效，请重新登录")
    return subject.removeprefix("admin:")


def user_item(db: Session, user: User) -> dict[str, Any]:
    phone = user.phone or phone_from_user_id(user.user_id) or ""
    post_count = (
        db.query(func.count(Post.id))
        .filter(Post.user_id == user.user_id, Post.is_deleted.is_(False))
        .scalar()
        or 0
    )
    comment_count = (
        db.query(func.count(Comment.id))
        .filter(Comment.user_id == user.user_id, Comment.is_deleted.is_(False))
        .scalar()
        or 0
    )
    likes_received = (
        db.query(func.coalesce(func.sum(Post.like_count), 0))
        .filter(Post.user_id == user.user_id, Post.is_deleted.is_(False))
        .scalar()
        or 0
    )
    return {
        "userId": user.user_id,
        "nickname": user.nickname or "即闪用户",
        "avatar": user.avatar or "",
        "phone": phone,
        "newPhone": user.new_phone or "",
        "clientType": user.client_type or "",
        "clientSubtype": user.client_subtype or "",
        "status": "normal" if user.is_active else "banned",
        "regTime": format_time(user.create_time),
        "postCount": post_count,
        "commentCount": comment_count,
        "likesReceived": likes_received,
        "bio": user.bio or "",
        "signature": user.bio or "",
        "gender": user.gender or "保密",
    }


def ensure_aware_utc(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def user_is_online(user: User, now: datetime | None = None) -> bool:
    last_time = ensure_aware_utc(user.last_time)
    if last_time is None or not user.is_active:
        return False
    current = ensure_aware_utc(now or utc_now()) or utc_now()
    return last_time >= current - timedelta(seconds=USER_ONLINE_WINDOW_SECONDS)


def user_online_item(user: User, now: datetime | None = None) -> dict[str, Any]:
    phone = user.phone or phone_from_user_id(user.user_id) or ""
    is_online = user_is_online(user, now)
    last_time = ensure_aware_utc(user.last_time)
    current = ensure_aware_utc(now or utc_now()) or utc_now()
    inactive_seconds = int((current - last_time).total_seconds()) if last_time else None
    return {
        "userId": user.user_id,
        "nickname": user.nickname or "即闪用户",
        "avatar": user.avatar or "",
        "phone": phone,
        "newPhone": user.new_phone or "",
        "clientType": user.client_type or "",
        "clientSubtype": user.client_subtype or "",
        "accountStatus": "normal" if user.is_active else "banned",
        "isOnline": is_online,
        "onlineStatus": "online" if is_online else "offline",
        "onlineStatusText": "在线" if is_online else "离线",
        "lastActiveAt": format_time(user.last_time),
        "lastTime": format_time(user.last_time),
        "inactiveSeconds": inactive_seconds,
        "regTime": format_time(user.create_time),
        "createdAt": format_time(user.create_time),
    }


USER_EXPORT_HEADERS = [
    ("用户ID", "userId"),
    ("昵称", "nickname"),
    ("手机号", "phone"),
    ("移动端类型", "clientType"),
    ("小程序类型", "clientSubtype"),
    ("openid", "openid"),
    ("unionid", "unionid"),
    ("性别", "gender"),
    ("个性签名", "bio"),
    ("省份", "province"),
    ("城市", "city"),
    ("区县", "district"),
    ("账号状态", "statusText"),
    ("注册时间", "createdAt"),
    ("更新时间", "updatedAt"),
]
USER_IMPORT_HEADER_ALIASES = {
    "用户ID": "userId",
    "用户Id": "userId",
    "userId": "userId",
    "user_id": "userId",
    "openid": "openid",
    "unionid": "unionid",
    "手机号": "phone",
    "手机号码": "phone",
    "phone": "phone",
    "移动端类型": "clientType",
    "客户端类型": "clientType",
    "clientType": "clientType",
    "client_type": "clientType",
    "platform": "clientType",
    "小程序类型": "clientSubtype",
    "小程序端类型": "clientSubtype",
    "clientSubtype": "clientSubtype",
    "client_subtype": "clientSubtype",
    "miniProgramType": "clientSubtype",
    "mpType": "clientSubtype",
    "昵称": "nickname",
    "用户昵称": "nickname",
    "nickname": "nickname",
    "头像": "avatar",
    "avatar": "avatar",
    "性别": "gender",
    "gender": "gender",
    "个性签名": "bio",
    "个人简介": "bio",
    "简介": "bio",
    "bio": "bio",
    "signature": "bio",
    "intro": "bio",
    "省份": "province",
    "province": "province",
    "城市": "city",
    "city": "city",
    "区县": "district",
    "地区": "district",
    "district": "district",
    "账号状态": "status",
    "状态": "status",
    "status": "status",
    "是否启用": "isActive",
    "isActive": "isActive",
}
USER_IMPORT_FIELDS = {"userId", "openid", "unionid", "phone", "clientType", "clientSubtype", "nickname", "avatar", "gender", "bio", "province", "city", "district", "status", "isActive"}


def export_user_row(user: User) -> dict[str, str]:
    phone = user.phone or phone_from_user_id(user.user_id) or ""
    return {
        "userId": user.user_id,
        "nickname": user.nickname or "",
        "phone": phone,
        "clientType": user.client_type or "",
        "clientSubtype": user.client_subtype or "",
        "openid": user.openid or "",
        "unionid": user.unionid or "",
        "gender": user.gender or "",
        "bio": user.bio or "",
        "province": user.province or "",
        "city": user.city or "",
        "district": user.district or "",
        "statusText": "正常" if user.is_active else "禁用",
        "createdAt": format_time(user.create_time),
        "updatedAt": format_time(user.update_time),
    }


def normalize_excel_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_user_status(value: str) -> bool | None:
    if not value:
        return None
    lowered = value.strip().lower()
    if lowered in {"normal", "active", "enabled", "true", "1", "正常", "启用", "已启用"}:
        return True
    if lowered in {"banned", "disabled", "false", "0", "禁用", "已禁用", "封禁"}:
        return False
    raise ValueError("账号状态只能填写 正常/禁用 或 normal/banned")


def read_user_import_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = FilePath(filename).suffix.lower()
    if suffix not in {".xls", ".xlsx"}:
        raise fail(status.HTTP_400_BAD_REQUEST, "导入文件格式不正确，仅支持 .xls 或 .xlsx")
    if not content:
        raise fail(status.HTTP_400_BAD_REQUEST, "导入文件不能为空")
    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        else:
            import xlrd

            workbook = xlrd.open_workbook(file_contents=content)
            sheet = workbook.sheet_by_index(0)
            raw_rows = [sheet.row_values(index) for index in range(sheet.nrows)]
    except Exception as exc:
        raise fail(status.HTTP_400_BAD_REQUEST, "文件格式无法解析，请确认是有效的 .xls 或 .xlsx 文件") from exc

    raw_rows = [row for row in raw_rows if any(normalize_excel_value(cell) for cell in row)]
    if not raw_rows:
        raise fail(status.HTTP_400_BAD_REQUEST, "导入文件没有可读取的数据")
    header_values = [normalize_excel_value(cell) for cell in raw_rows[0]]
    fields = [USER_IMPORT_HEADER_ALIASES.get(header) for header in header_values]
    if not any(field in USER_IMPORT_FIELDS for field in fields):
        raise fail(status.HTTP_400_BAD_REQUEST, "表头格式不正确，请包含 用户ID、昵称、手机号、账号状态 等字段")

    rows: list[dict[str, str]] = []
    for raw_row in raw_rows[1:]:
        item: dict[str, str] = {}
        for index, field in enumerate(fields):
            if not field or field not in USER_IMPORT_FIELDS:
                continue
            item[field] = normalize_excel_value(raw_row[index] if index < len(raw_row) else None)
        if any(item.values()):
            rows.append(item)
    if not rows:
        raise fail(status.HTTP_400_BAD_REQUEST, "导入文件没有有效用户数据")
    return rows


def export_users_xlsx(users: list[User]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用户列表"
    sheet.append([header for header, _ in USER_EXPORT_HEADERS])
    for user in users:
        data = export_user_row(user)
        sheet.append([data.get(field, "") for _, field in USER_EXPORT_HEADERS])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_users_xls(users: list[User]) -> bytes:
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("用户列表")
    for column, (header, _) in enumerate(USER_EXPORT_HEADERS):
        sheet.write(0, column, header)
    for row_index, user in enumerate(users, start=1):
        data = export_user_row(user)
        for column, (_, field) in enumerate(USER_EXPORT_HEADERS):
            sheet.write(row_index, column, data.get(field, ""))
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def user_export_response(users: list[User], format: str) -> StreamingResponse:
    if format == "xlsx":
        content = export_users_xlsx(users)
        filename = "instant-flash-users.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = export_users_xls(users)
        filename = "instant-flash-users.xls"
        media_type = "application/vnd.ms-excel"
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def import_users_from_rows(db: Session, rows: list[dict[str, str]]) -> dict[str, Any]:
    errors: list[dict[str, Any]] = []
    created = 0
    updated = 0
    for index, row in enumerate(rows, start=2):
        user_id = row.get("userId", "")
        phone = normalize_phone(row.get("phone", "")) or phone_from_user_id(user_id) or ""
        openid = row.get("openid", "")
        unionid = row.get("unionid", "")
        if not any([user_id, phone, openid, unionid]):
            errors.append({"row": index, "message": "至少填写 用户ID、手机号、openid、unionid 中的一个"})
            continue
        try:
            db.flush()
            conditions = []
            if user_id:
                conditions.append(User.user_id == user_id)
            if phone:
                conditions.append(User.phone == phone)
            if openid:
                conditions.append(User.openid == openid)
            if unionid:
                conditions.append(User.unionid == unionid)
            matches = db.query(User).filter(or_(*conditions)).all()
            if len({user.id for user in matches}) > 1:
                errors.append({"row": index, "message": "用户ID/手机号/openid/unionid 命中多个用户，请检查唯一字段"})
                continue
            user = matches[0] if matches else None
            if user is None:
                user = User(user_id=user_id or new_business_id("usr"))
                db.add(user)
                created += 1
            else:
                updated += 1

            if phone:
                user.phone = phone
            client_type = normalize_client_type(row.get("clientType"))
            client_subtype = normalize_client_subtype(row.get("clientSubtype"))
            if client_type:
                user.client_type = client_type
            if client_subtype:
                user.client_subtype = client_subtype
            for field in ("openid", "unionid", "nickname", "avatar", "gender", "bio", "province", "city", "district"):
                value = row.get(field)
                if value:
                    setattr(user, field, value)
            status_value = row.get("status", "") or row.get("isActive", "")
            active = normalize_user_status(status_value)
            if active is not None:
                user.is_active = active
            user.last_time = utc_now()
        except ValueError as exc:
            errors.append({"row": index, "message": str(exc)})
        except Exception as exc:
            errors.append({"row": index, "message": f"导入失败：{exc}"})

    if errors:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": 400, "message": "导入失败，请检查错误行", "data": {"errors": errors}},
        )
    db.commit()
    return {"created": created, "updated": updated, "total": created + updated, "errors": []}


def post_status(post: Post) -> str:
    if post.status in {"offline", "draft"}:
        return post.status
    return "online"


def post_item(post: Post) -> dict[str, Any]:
    author = post.author
    media = post.images or []
    videos = [item for item in media if media_type(item) == "video"]
    return {
        "postId": post.post_id,
        "userId": post.user_id,
        "nickname": author.nickname if author and author.nickname else "即闪用户",
        "avatar": author.avatar if author and author.avatar else "",
        "content": post.content,
        "images": media,
        "videos": videos,
        "media": media,
        "topics": post.topics or [],
        "location": post.location or "",
        "province": post.province or "",
        "city": post.city or "",
        "district": post.district or "",
        "visibility": post.visibility,
        "likes": post.like_count,
        "comments": post.comment_count,
        "shares": post.share_count,
        "status": post_status(post),
        "pubTime": format_time(post.create_time),
    }


def comment_item(db: Session, comment: Comment) -> dict[str, Any]:
    user = db.query(User).filter(User.user_id == comment.user_id).one_or_none()
    reply_to = (
        db.query(User).filter(User.user_id == comment.reply_to_user_id).one_or_none()
        if comment.reply_to_user_id
        else None
    )
    return {
        "commentId": comment.comment_id,
        "postId": comment.post_id,
        "userId": comment.user_id,
        "nickname": user.nickname if user and user.nickname else "即闪用户",
        "avatar": user.avatar if user and user.avatar else "",
        "content": comment.content,
        "parentId": comment.parent_id,
        "replyToUserId": comment.reply_to_user_id,
        "replyToNickname": reply_to.nickname if reply_to else None,
        "pubTime": format_time(comment.create_time),
    }


def like_item(db: Session, like: PostLike) -> dict[str, Any]:
    user = db.query(User).filter(User.user_id == like.user_id).one_or_none()
    post = db.query(Post).filter(Post.post_id == like.post_id).one_or_none()
    return {
        "likeId": str(like.id),
        "postId": like.post_id,
        "userId": like.user_id,
        "nickname": user.nickname if user and user.nickname else "即闪用户",
        "avatar": user.avatar if user and user.avatar else "",
        "content": post.content if post else "",
        "postStatus": post_status(post) if post else "",
        "likedAt": format_time(like.create_time),
        "createdAt": format_time(like.create_time),
    }


def share_item(db: Session, share: PostShare) -> dict[str, Any]:
    user = db.query(User).filter(User.user_id == share.user_id).one_or_none()
    post = db.query(Post).filter(Post.post_id == share.post_id).one_or_none()
    return {
        "shareId": str(share.id),
        "postId": share.post_id,
        "userId": share.user_id,
        "nickname": user.nickname if user and user.nickname else "即闪用户",
        "avatar": user.avatar if user and user.avatar else "",
        "content": post.content if post else "",
        "postStatus": post_status(post) if post else "",
        "scene": share.scene or "",
        "platform": share.platform or "",
        "sharedAt": format_time(share.create_time),
        "createdAt": format_time(share.create_time),
    }


SCAM_KEYWORDS = ("骗", "诈骗", "杀猪盘", "套路", "上当", "虚假", "刷单", "被扣", "举报")


def tag_search_count(name: str) -> int:
    value = 0
    for index, char in enumerate(name):
        value += ord(char) * (index + 3)
    return (value % 8500) + 1200


def tag_metrics(db: Session, name: str) -> dict[str, int]:
    tag_text = f"#{name}"
    post_query = db.query(Post).filter(
        Post.is_deleted.is_(False),
        Post.content.ilike(f"%{tag_text}%"),
    )
    post_count = post_query.count()
    scam_count = post_query.filter(or_(*(Post.content.ilike(f"%{keyword}%") for keyword in SCAM_KEYWORDS))).count()
    return {
        "postCount": post_count,
        "associatedPostCount": post_count,
        "linkedPostCount": post_count,
        "scamCount": scam_count,
        "fraudCount": scam_count,
        "searchCount": tag_search_count(name),
    }


def tag_item(db: Session, tag: AdminTag) -> dict[str, Any]:
    return {
        "tagId": tag.tag_id,
        "name": tag.name,
        "color": tag.color or "",
        "sort": tag.sort,
        "status": tag.status,
        "remark": tag.remark or "",
        **tag_metrics(db, tag.name),
        "createdAt": format_time(tag.create_time),
        "updatedAt": format_time(tag.update_time),
    }


def region_item(region: AdminRegion) -> dict[str, Any]:
    return {
        "regionId": region.region_id,
        "parentId": region.parent_id,
        "name": region.name,
        "code": region.code,
        "level": region.level,
        "sort": region.sort,
        "status": region.status,
        "createdAt": format_time(region.create_time),
        "updatedAt": format_time(region.update_time),
    }


def dictionary_item(dictionary: AdminDictionary) -> dict[str, Any]:
    return {
        "dictId": dictionary.dict_id,
        "parentId": dictionary.parent_id,
        "type": dictionary.type,
        "label": dictionary.label,
        "value": dictionary.value,
        "sort": dictionary.sort,
        "status": dictionary.status,
        "remark": dictionary.remark or "",
        "createdAt": format_time(dictionary.create_time),
        "updatedAt": format_time(dictionary.update_time),
    }


def dictionary_tree_items(dictionaries: list[AdminDictionary]) -> list[dict[str, Any]]:
    children_by_parent: dict[str | None, list[AdminDictionary]] = {}
    ids = {dictionary.dict_id for dictionary in dictionaries}
    for dictionary in dictionaries:
        parent_id = dictionary.parent_id if dictionary.parent_id in ids else None
        children_by_parent.setdefault(parent_id, []).append(dictionary)

    for children in children_by_parent.values():
        children.sort(key=lambda item: (item.type, item.sort, item.create_time), reverse=False)

    def build(parent_id: str | None) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for dictionary in children_by_parent.get(parent_id, []):
            item = dictionary_item(dictionary)
            item["children"] = build(dictionary.dict_id)
            item["childCount"] = len(item["children"])
            result.append(item)
        return result

    return build(None)


def get_dictionary_or_404(db: Session, dict_id: str) -> AdminDictionary:
    dictionary = db.query(AdminDictionary).filter(AdminDictionary.dict_id == dict_id).one_or_none()
    if dictionary is None:
        raise fail(status.HTTP_404_NOT_FOUND, "字典未找到")
    return dictionary


def validate_dictionary_parent(db: Session, parent_id: str | None, dictionary_type: str, self_id: str | None = None) -> None:
    if not parent_id:
        return
    if parent_id == self_id:
        raise fail(status.HTTP_400_BAD_REQUEST, "上级字典不能选择自己")
    parent = get_dictionary_or_404(db, parent_id)
    if parent.type != dictionary_type:
        raise fail(status.HTTP_400_BAD_REQUEST, "上级字典类型必须和当前字典类型一致")
    cursor = parent
    while cursor.parent_id:
        if cursor.parent_id == self_id:
            raise fail(status.HTTP_400_BAD_REQUEST, "上级字典不能选择自己的下级")
        cursor = get_dictionary_or_404(db, cursor.parent_id)


def dictionary_scope_query(db: Session, dictionary_type: str, parent_id: str | None):
    query = db.query(AdminDictionary).filter(AdminDictionary.type == dictionary_type)
    if parent_id:
        return query.filter(AdminDictionary.parent_id == parent_id)
    return query.filter(AdminDictionary.parent_id.is_(None))


def validate_dictionary_unique(
    db: Session,
    dictionary_type: str,
    parent_id: str | None,
    label: str,
    value: str,
    self_id: str | None = None,
) -> None:
    query = dictionary_scope_query(db, dictionary_type, parent_id)
    if self_id:
        query = query.filter(AdminDictionary.dict_id != self_id)
    duplicates = query.all()
    if any(item.label == label for item in duplicates):
        raise fail(status.HTTP_400_BAD_REQUEST, "同一上级下字典标签已存在")
    if any(item.value == value for item in duplicates):
        raise fail(status.HTTP_400_BAD_REQUEST, "同一上级下字典键值已存在")


def system_message_item(message: AdminSystemMessage) -> dict[str, Any]:
    return {
        "messageId": message.message_id,
        "title": message.title,
        "content": message.content,
        "type": message.type,
        "target": message.target,
        "status": message.status,
        "isPinned": message.is_pinned,
        "createdAt": format_time(message.create_time),
        "updatedAt": format_time(message.update_time),
    }


def feedback_field_item(field: AdminFeedbackField) -> dict[str, Any]:
    return {
        "fieldId": field.field_id,
        "formId": field.form_id,
        "fieldKey": field.field_key,
        "label": field.label,
        "type": field.type,
        "placeholder": field.placeholder or "",
        "required": field.required,
        "options": field.options or [],
        "sort": field.sort,
        "status": field.status,
        "isDefault": field.is_default,
        "remark": field.remark or "",
        "createdAt": format_time(field.create_time),
        "updatedAt": format_time(field.update_time),
    }


def feedback_config_item(db: Session, config: AdminFeedbackFormConfig) -> dict[str, Any]:
    fields = (
        db.query(AdminFeedbackField)
        .filter(AdminFeedbackField.form_id == config.config_id)
        .order_by(AdminFeedbackField.sort.asc(), AdminFeedbackField.create_time.asc())
        .all()
    )
    active_fields = [field for field in fields if field.status == "enabled"]
    return {
        "configId": config.config_id,
        "title": config.title,
        "menuTitle": config.menu_title,
        "description": config.description or "",
        "submitButtonText": config.submit_button_text,
        "successMessage": config.success_message,
        "status": config.status,
        "remark": config.remark or "",
        "fields": [feedback_field_item(field) for field in fields],
        "activeFields": [feedback_field_item(field) for field in active_fields],
        "createdAt": format_time(config.create_time),
        "updatedAt": format_time(config.update_time),
    }


def feedback_submission_item(feedback: FeedbackSubmission) -> dict[str, Any]:
    return {
        "feedbackId": feedback.feedback_id,
        "userId": feedback.user_id or "",
        "phone": feedback.phone or "",
        "title": feedback.title or "",
        "content": feedback.content,
        "payload": feedback.payload or {},
        "status": feedback.status,
        "reply": feedback.reply or "",
        "ip": feedback.ip or "",
        "userAgent": feedback.user_agent or "",
        "remark": feedback.remark or "",
        "createdAt": format_time(feedback.create_time),
        "updatedAt": format_time(feedback.update_time),
    }


def announcement_item(announcement: AdminAnnouncement) -> dict[str, Any]:
    return {
        "id": announcement.announcement_id,
        "title": announcement.title,
        "type": announcement.type,
        "content": announcement.content,
        "link": announcement.link or "",
        "pinned": announcement.pinned,
        "startTime": announcement.start_time or "",
        "endTime": announcement.end_time or "",
        "status": announcement.status,
        "createdAt": format_time(announcement.create_time),
        "updatedAt": format_time(announcement.update_time),
    }


def version_item(version: AdminVersion) -> dict[str, Any]:
    return {
        "id": version.version_id,
        "platform": version.platform,
        "version": version.version,
        "build": version.build,
        "forceUpdate": version.force_update,
        "status": version.status,
        "betaPct": version.beta_pct,
        "notes": version.notes,
        "notesType": version.notes_type,
        "downloadUrl": version.download_url or "",
        "releaseTime": version.release_time or format_time(version.create_time),
        "createdAt": format_time(version.create_time),
        "updatedAt": format_time(version.update_time),
    }


def file_size_text(size_bytes: int) -> str:
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    if size_bytes >= 1024:
        return f"{size_bytes / 1024:.2f} KB"
    return f"{size_bytes} B"


def package_item(package: AdminPackage) -> dict[str, Any]:
    return {
        "packageId": package.package_id,
        "platform": package.platform,
        "version": package.version,
        "build": package.build or "",
        "displayName": package.display_name,
        "fileName": package.original_filename,
        "originalFileName": package.original_filename,
        "filePath": package.file_path,
        "downloadUrl": package.download_url,
        "md5": package.md5,
        "sizeBytes": package.size_bytes,
        "sizeText": file_size_text(package.size_bytes),
        "status": package.status,
        "remark": package.remark or "",
        "uploadTime": format_time(package.create_time),
        "createdAt": format_time(package.create_time),
        "updatedAt": format_time(package.update_time),
    }


def package_version_summary_item(row: tuple[str, int, int | None, datetime | None]) -> dict[str, Any]:
    version, total, total_size, last_upload = row
    return {
        "version": version,
        "total": total,
        "packageCount": total,
        "sizeBytes": total_size or 0,
        "sizeText": file_size_text(total_size or 0),
        "lastUploadTime": format_time(last_upload),
    }


def account_item(account: AdminAccount) -> dict[str, Any]:
    return {
        "accountId": account.account_id,
        "account_id": account.account_id,
        "username": account.username,
        "nickname": account.nickname,
        "avatar": account.avatar or "",
        "role": account.role,
        "permissions": account.permissions or [],
        "status": account.status,
        "email": account.email or "",
        "phone": account.phone or "",
        "createTime": format_time(account.create_time),
        "lastLogin": account.last_login or format_time(account.last_time),
        "remark": account.remark or "",
    }


def role_item(role: AdminRole) -> dict[str, Any]:
    return {
        "roleId": role.role_id,
        "roleKey": role.role_key,
        "label": role.label,
        "icon": role.icon or "",
        "permissions": role.permissions or [],
        "sort": role.sort,
        "status": role.status,
        "isDefault": role.is_default,
        "remark": role.remark or "",
        "createdAt": format_time(role.create_time),
        "updatedAt": format_time(role.update_time),
    }


def permission_item(permission: AdminPermission) -> dict[str, Any]:
    return {
        "permissionId": permission.permission_id,
        "permissionKey": permission.permission_key,
        "label": permission.label,
        "description": permission.description or "",
        "sort": permission.sort,
        "status": permission.status,
        "isDefault": permission.is_default,
        "remark": permission.remark or "",
        "createdAt": format_time(permission.create_time),
        "updatedAt": format_time(permission.update_time),
    }


def operation_log_item(log: AdminOperationLog, current: bool = False) -> dict[str, Any]:
    return {
        "logId": log.log_id,
        "accountId": log.account_id or "",
        "username": log.username or "",
        "category": log.category,
        "action": log.action,
        "title": f"{log.title}（当前登录）" if current else log.title,
        "content": log.content or "",
        "status": log.status,
        "ip": log.ip or "",
        "location": log.location or "",
        "userAgent": log.user_agent or "",
        "createdAt": format_time(log.create_time),
        "time": format_time(log.create_time),
        "isCurrent": current,
    }


def security_setting_item(setting: AdminSecuritySetting, account: AdminAccount) -> dict[str, Any]:
    password_policy_passed = len(account.password or "") >= 6 and any(char.isdigit() for char in account.password or "")
    score = 55
    if setting.password_policy_enabled and password_policy_passed:
        score += 30
    if setting.mfa_enabled:
        score += 15
    score = min(score, 100)
    if score >= 90:
        level = "高"
        levelKey = "high"
    elif score >= 70:
        level = "中等"
        levelKey = "medium"
    else:
        level = "低"
        levelKey = "low"
    return {
        "accountId": account.account_id,
        "username": account.username,
        "score": score,
        "level": level,
        "levelKey": levelKey,
        "mfaEnabled": setting.mfa_enabled,
        "passwordPolicyEnabled": setting.password_policy_enabled,
        "passwordPolicyPassed": password_policy_passed,
        "mfaStatusText": "已开启。登录时需要二次验证。" if setting.mfa_enabled else "未开启。仅支持账号密码登录。",
        "passwordPolicyText": "已符合长度及混合字符要求。" if password_policy_passed else "密码强度较弱，建议包含数字和字母。",
        "recommendation": "建议定期修改登录密码并绑定多因子安全设备。",
        "remark": setting.remark or "",
        "updatedAt": format_time(setting.update_time),
    }


def access_rule_item(rule: AdminAccessRule) -> dict[str, Any]:
    return {
        "ruleId": rule.rule_id,
        "type": rule.rule_type,
        "typeText": "黑名单" if rule.rule_type == "blacklist" else "白名单",
        "ip": rule.ip or "",
        "method": rule.method or "ALL",
        "path": rule.path or "",
        "status": rule.status,
        "statusText": "启用" if rule.status == "enabled" else "禁用",
        "remark": rule.remark or "",
        "createdAt": format_time(rule.create_time),
        "updatedAt": format_time(rule.update_time),
    }


def menu_item(menu: AdminMenu) -> dict[str, Any]:
    return {
        "menuId": menu.menu_id,
        "parentId": menu.parent_id,
        "title": menu.title,
        "path": menu.path,
        "name": menu.name,
        "component": menu.component or "",
        "redirect": menu.redirect or "",
        "icon": menu.icon or "",
        "type": menu.type,
        "permission": menu.permission or "",
        "sort": menu.sort,
        "status": menu.status,
        "visible": menu.visible,
        "keepAlive": menu.keep_alive,
        "affix": menu.affix,
        "externalLink": menu.external_link or "",
        "remark": menu.remark or "",
        "meta": {
            "title": menu.title,
            "icon": menu.icon or "",
            "permission": menu.permission or "",
            "keepAlive": menu.keep_alive,
            "affix": menu.affix,
            "hidden": not menu.visible,
        },
        "createdAt": format_time(menu.create_time),
        "updatedAt": format_time(menu.update_time),
    }


def menu_tree_items(menus: list[AdminMenu]) -> list[dict[str, Any]]:
    ids = {menu.menu_id for menu in menus}
    children_by_parent: dict[str | None, list[AdminMenu]] = {}
    for menu in menus:
        parent_id = menu.parent_id if menu.parent_id in ids else None
        children_by_parent.setdefault(parent_id, []).append(menu)
    for children in children_by_parent.values():
        children.sort(key=lambda item: (item.sort, item.create_time))

    def build(parent_id: str | None) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for menu in children_by_parent.get(parent_id, []):
            item = menu_item(menu)
            item["children"] = build(menu.menu_id)
            item["childCount"] = len(item["children"])
            result.append(item)
        return result

    return build(None)


def permitted_menu_tree_items(menus: list[AdminMenu], account: AdminAccount, active_permissions: set[str]) -> list[dict[str, Any]]:
    allowed_permissions = set(account.permissions or []) & active_permissions
    is_superadmin = account.role == "superadmin"

    def can_access(menu: AdminMenu) -> bool:
        if menu.name in ADMIN_ONLY_ROUTE_NAMES and account.role not in ADMIN_ROUTE_ROLES:
            return False
        if not menu.permission:
            return True
        if menu.permission not in active_permissions:
            return False
        return is_superadmin or menu.permission in allowed_permissions

    all_tree = menu_tree_items([menu for menu in menus if menu.status == "enabled" and menu.type != "button"])

    def prune(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        menu_by_id = {menu.menu_id: menu for menu in menus}
        for node in nodes:
            children = prune(node.get("children", []))
            source = menu_by_id.get(node["menuId"])
            if source and (children or (source.type != "catalog" and can_access(source))):
                node["children"] = children
                node["childCount"] = len(children)
                result.append(node)
        return result

    return prune(all_tree)


def flatten_menu_tree(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for node in nodes:
        result.append({key: value for key, value in node.items() if key != "children"})
        result.extend(flatten_menu_tree(node.get("children", [])))
    return result


def notification_item(message: Message) -> dict[str, Any]:
    return {
        "messageId": message.message_id,
        "title": message.title or "",
        "content": message.content or "",
        "type": message.type,
        "isRead": message.is_read,
        "time": format_time(message.create_time),
        "createdAt": format_time(message.create_time),
        "updatedAt": format_time(message.update_time),
        "sourceId": message.post_id,
    }


def dashboard_trend_buckets(period: str) -> list[tuple[str, datetime, datetime]]:
    now = utc_now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if period == "today":
        points = [2, 6, 10, 14, 18, 22, 24]
        buckets: list[tuple[str, datetime, datetime]] = []
        previous = today_start
        for hour in points:
            current = today_start + timedelta(hours=hour)
            label = "今日汇总" if hour == 24 else f"{hour:02d}:00"
            bounded_end = min(current, now)
            if bounded_end < previous:
                bounded_end = previous
            buckets.append((label, previous, bounded_end))
            previous = current
        return buckets

    if period == "week":
        week_start = today_start - timedelta(days=today_start.weekday())
        labels = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        return [
            (
                labels[index],
                week_start + timedelta(days=index),
                max(week_start + timedelta(days=index), min(week_start + timedelta(days=index + 1), now)),
            )
            for index in range(7)
        ]

    month_start = today_start.replace(day=1)
    buckets = []
    cursor = month_start
    while cursor <= now:
        end = cursor + timedelta(days=1)
        buckets.append((f"{cursor.day:02d}日", cursor, min(end, now)))
        cursor = end
    return buckets


def count_between(db: Session, model: type[Any], field: Any, start: datetime, end: datetime, *conditions: Any) -> int:
    query = db.query(func.count(model.id)).filter(field >= start, field < end)
    for condition in conditions:
        query = query.filter(condition)
    return query.scalar() or 0


def distinct_users_between(db: Session, start: datetime, end: datetime) -> int:
    user_ids: set[str] = set()
    for value in db.query(Post.user_id).filter(Post.create_time >= start, Post.create_time < end, Post.is_deleted.is_(False)).all():
        user_ids.add(value[0])
    for value in db.query(Comment.user_id).filter(Comment.create_time >= start, Comment.create_time < end, Comment.is_deleted.is_(False)).all():
        user_ids.add(value[0])
    for value in db.query(PostLike.user_id).filter(PostLike.create_time >= start, PostLike.create_time < end).all():
        user_ids.add(value[0])
    for value in db.query(PostShare.user_id).filter(PostShare.create_time >= start, PostShare.create_time < end).all():
        user_ids.add(value[0])
    for value in db.query(Message.user_id).filter(Message.create_time >= start, Message.create_time < end).all():
        user_ids.add(value[0])
    return len(user_ids)


def ensure_admin_notification(db: Session, admin_subject: str, message: AdminSystemMessage) -> None:
    admin_user_id = f"admin:{admin_subject}"
    exists = (
        db.query(Message)
        .filter(Message.user_id == admin_user_id, Message.type == "admin_system", Message.post_id == message.message_id)
        .one_or_none()
    )
    if exists is not None:
        exists.title = message.title
        exists.content = message.content
        exists.last_time = utc_now()
        return
    db.add(
        Message(
            message_id=new_business_id("ntf"),
            user_id=admin_user_id,
            sender_id=None,
            type="admin_system",
            title=message.title,
            content=message.content,
            post_id=message.message_id,
            is_read=False,
        )
    )


def publish_system_message_to_users(db: Session, message: AdminSystemMessage) -> int:
    query = db.query(User).filter(User.is_active.is_(True))
    if message.target == "new":
        query = query.order_by(User.create_time.desc()).limit(200)
    users = query.all()
    created = 0
    for user in users:
        exists = (
            db.query(Message)
            .filter(Message.user_id == user.user_id, Message.type == "system", Message.post_id == message.message_id)
            .one_or_none()
        )
        if exists is not None:
            exists.title = message.title
            exists.content = message.content
            exists.last_time = utc_now()
            continue
        db.add(
            Message(
                message_id=new_business_id("ntf"),
                user_id=user.user_id,
                sender_id=None,
                type="system",
                title=message.title,
                content=message.content,
                post_id=message.message_id,
                is_read=False,
            )
        )
        created += 1
    return created


def get_or_create_agreement(db: Session, agreement_type: str) -> AdminAgreement:
    agreement = db.query(AdminAgreement).filter(AdminAgreement.type == agreement_type).one_or_none()
    if agreement is not None:
        return agreement
    agreement = AdminAgreement(
        type=agreement_type,
        content=DEFAULT_AGREEMENTS[agreement_type],
    )
    db.add(agreement)
    db.commit()
    db.refresh(agreement)
    return agreement


@router.post(
    "/auth/login",
    response_model=AdminResponse,
    summary="后台登录",
    description="后台管理系统登录接口。演示账号 admin，密码 123456。",
)
def admin_login(payload: AdminLoginRequest, db: Annotated[Session, Depends(get_db)], request: Request) -> dict[str, Any]:
    seed_accounts_if_empty(db)
    seed_permissions_if_empty(db)
    account = db.query(AdminAccount).filter(AdminAccount.username == payload.username).one_or_none()
    if account is None or account.password != payload.password:
        write_admin_log(
            db,
            account=account,
            username=payload.username,
            category="login",
            action="login_failed",
            title="密码校验错误警告",
            content="用户名或密码错误",
            status_value="warning",
            request=request,
        )
        db.commit()
        raise fail(status.HTTP_400_BAD_REQUEST, "用户名或密码错误")
    if account.status != "active":
        write_admin_log(
            db,
            account=account,
            username=payload.username,
            category="login",
            action="login_disabled",
            title="禁用账号登录拦截",
            content="账号已禁用",
            status_value="warning",
            request=request,
        )
        db.commit()
        raise fail(status.HTTP_403_FORBIDDEN, "账号已禁用，请联系管理员")
    account.last_login = format_time(utc_now())
    account.last_time = utc_now()
    write_admin_log(
        db,
        account=account,
        username=account.username,
        category="login",
        action="login_success",
        title="登录系统成功",
        content="后台账号登录成功",
        status_value="success",
        request=request,
    )
    db.commit()
    token = create_access_token(f"admin:{account.username}", token_type="admin")
    return ok(
        {
            "token": token,
            "username": account.username,
            "nickname": account.nickname,
            "role": account.role,
            "permissions": account_effective_permissions(db, account),
        },
        "登录成功",
    )


@router.post(
    "/auth/logout",
    response_model=AdminResponse,
    summary="后台退出登录",
    description="PC 后台退出登录接口。后端会删除 Redis 中的 Token 登录状态，前端同时清理本地 Token。",
)
def admin_logout(
    _: Annotated[str, Depends(get_admin_subject)],
    credentials: Annotated[HTTPAuthorizationCredentials | None, Depends(admin_bearer)],
) -> dict[str, Any]:
    if credentials is not None:
        revoke_access_token(credentials.credentials)
    return ok({}, "退出成功")


@router.get(
    "/dashboard/metrics",
    response_model=AdminResponse,
    summary="看板指标",
    description="获取后台首页数据看板指标，包括用户、内容、评论和点赞统计。",
)
def dashboard_metrics(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    total_users = db.query(func.count(User.id)).scalar() or 0
    active_users = db.query(func.count(User.id)).filter(User.is_active.is_(True)).scalar() or 0
    post_query = db.query(Post).filter(Post.is_deleted.is_(False))
    total_posts = post_query.count()
    offline_posts = post_query.filter(Post.status == "offline").count()
    online_posts = total_posts - offline_posts
    total_comments = db.query(func.count(Comment.id)).filter(Comment.is_deleted.is_(False)).scalar() or 0
    total_likes = (
        db.query(func.coalesce(func.sum(Post.like_count), 0))
        .filter(Post.is_deleted.is_(False))
        .scalar()
        or 0
    )
    return ok(
        {
            "totalUsers": total_users,
            "activeUsers": active_users,
            "totalPosts": total_posts,
            "onlinePosts": online_posts,
            "offlinePosts": offline_posts,
            "totalComments": total_comments,
            "totalLikes": total_likes,
        }
    )


@router.get(
    "/dashboard/trends",
    response_model=AdminResponse,
    summary="看板趋势图",
    description="获取后台首页趋势图数据，支持流量与内容、用户活跃与增长，以及今日/本周/本月。",
)
def dashboard_trends(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    type: Annotated[str, Query(pattern="^(traffic_content|user_growth)$", description="趋势类型：traffic_content 流量与内容，user_growth 用户活跃与增长")] = "traffic_content",
    period: Annotated[str, Query(pattern="^(today|week|month)$", description="时间范围：today 今日，week 本周，month 本月")] = "today",
) -> dict[str, Any]:
    buckets = dashboard_trend_buckets(period)
    labels: list[str] = []
    visits: list[int] = []
    posts: list[int] = []
    active_users: list[int] = []
    new_users: list[int] = []

    for label, start, end in buckets:
        labels.append(label)
        post_count = count_between(db, Post, Post.create_time, start, end, Post.is_deleted.is_(False))
        comment_count = count_between(db, Comment, Comment.create_time, start, end, Comment.is_deleted.is_(False))
        like_count = count_between(db, PostLike, PostLike.create_time, start, end)
        share_count = count_between(db, PostShare, PostShare.create_time, start, end)
        message_count = count_between(db, Message, Message.create_time, start, end)
        visits.append(post_count + comment_count + like_count + share_count + message_count)
        posts.append(post_count)
        active_users.append(distinct_users_between(db, start, end))
        new_users.append(count_between(db, User, User.create_time, start, end))

    if type == "user_growth":
        series = [
            {"key": "activeUsers", "name": "活跃用户趋势 (Active Users)", "type": "line", "data": active_users},
            {"key": "newUsers", "name": "新增用户统计 (New Users)", "type": "bar", "data": new_users},
        ]
    else:
        series = [
            {"key": "visits", "name": "活跃流量趋势 (Visits)", "type": "line", "data": visits},
            {"key": "posts", "name": "内容发布量统计 (Posts)", "type": "bar", "data": posts},
        ]

    return ok(
        {
            "type": type,
            "period": period,
            "labels": labels,
            "series": series,
            "visits": visits,
            "posts": posts,
            "activeUsers": active_users,
            "newUsers": new_users,
            "summary": {
                "visits": sum(visits),
                "posts": sum(posts),
                "activeUsers": sum(active_users),
                "newUsers": sum(new_users),
            },
        }
    )


@router.get(
    "/security/overview",
    response_model=AdminResponse,
    summary="账号安全概览",
    description="账号安全页概览接口，返回安全评分、二次验证、密码策略和最近 3 次登录日志。",
)
def admin_security_overview(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    account = get_current_admin_account(db, admin_subject)
    setting = get_or_create_security_setting(db, account)
    logs = (
        db.query(AdminOperationLog)
        .filter(AdminOperationLog.account_id == account.account_id, AdminOperationLog.category == "login")
        .order_by(AdminOperationLog.create_time.desc())
        .limit(3)
        .all()
    )
    return ok({**security_setting_item(setting, account), "recentLoginLogs": [operation_log_item(log, index == 0) for index, log in enumerate(logs)]})


@router.get(
    "/security/settings",
    response_model=AdminResponse,
    summary="账号安全设置",
    description="查询当前后台账号安全设置。",
)
def get_admin_security_settings(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    account = get_current_admin_account(db, admin_subject)
    return ok(security_setting_item(get_or_create_security_setting(db, account), account))


@router.put(
    "/security/settings",
    response_model=AdminResponse,
    summary="修改账号安全设置",
    description="修改当前后台账号二次登录验证和密码策略审核状态。",
)
def update_admin_security_settings(
    payload: AdminSecuritySettingPayload,
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
    request: Request,
) -> dict[str, Any]:
    account = get_current_admin_account(db, admin_subject)
    setting = get_or_create_security_setting(db, account)
    if payload.mfaEnabled is not None:
        setting.mfa_enabled = payload.mfaEnabled
    if payload.passwordPolicyEnabled is not None:
        setting.password_policy_enabled = payload.passwordPolicyEnabled
    if payload.remark is not None:
        setting.remark = payload.remark
    setting.last_time = utc_now()
    write_admin_log(
        db,
        account=account,
        username=account.username,
        category="security",
        action="security_settings_update",
        title="账号安全设置更新",
        content="修改二次登录验证或密码策略审核设置",
        status_value="success",
        request=request,
    )
    db.commit()
    db.refresh(setting)
    return ok(security_setting_item(setting, account), "安全设置更新成功")


@router.get(
    "/security/login-logs",
    response_model=AdminResponse,
    summary="当前账号登录日志",
    description="查询当前后台账号登录日志，安全设置页可用于最近登录记录展示。",
)
def list_current_admin_login_logs(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    account = get_current_admin_account(db, admin_subject)
    query = db.query(AdminOperationLog).filter(AdminOperationLog.account_id == account.account_id, AdminOperationLog.category == "login")
    total = query.count()
    logs = query.order_by(AdminOperationLog.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [operation_log_item(log, index == 0 and page == 1) for index, log in enumerate(logs)], "total": total})


@router.get(
    "/access-rules",
    response_model=AdminResponse,
    summary="黑白名单列表",
    description="接口限流黑名单/白名单列表。白名单命中跳过限流，黑名单命中直接拦截。",
)
def list_admin_access_rules(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="IP、路径、备注关键词")] = None,
    type: Annotated[str | None, Query(pattern="^(blacklist|whitelist)$", description="名单类型")] = None,
    status_filter: Annotated[str | None, Query(alias="status", pattern="^(enabled|disabled)$", description="状态")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=200, description="每页数量")] = 20,
) -> dict[str, Any]:
    query = db.query(AdminAccessRule)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter(
            or_(
                AdminAccessRule.ip.ilike(like_keyword),
                AdminAccessRule.path.ilike(like_keyword),
                AdminAccessRule.method.ilike(like_keyword),
                AdminAccessRule.remark.ilike(like_keyword),
            )
        )
    if type:
        query = query.filter(AdminAccessRule.rule_type == type)
    if status_filter:
        query = query.filter(AdminAccessRule.status == status_filter)
    total = query.count()
    rules = query.order_by(AdminAccessRule.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [access_rule_item(rule) for rule in rules], "total": total})


@router.post(
    "/access-rules",
    response_model=AdminResponse,
    summary="新增黑白名单",
    description="新增接口限流黑名单或白名单规则。",
)
def create_admin_access_rule(
    payload: AdminAccessRulePayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    rule = AdminAccessRule(
        rule_id=new_business_id("rule"),
        rule_type=payload.type,
        ip=(payload.ip or "").strip() or None,
        method=(payload.method or "ALL").upper(),
        path=(payload.path or "").strip() or None,
        status=payload.status,
        remark=payload.remark or "",
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return ok(access_rule_item(rule), "黑白名单规则创建成功")


@router.get("/access-rules/{ruleId}", response_model=AdminResponse, summary="黑白名单详情", description="根据规则 ID 查询黑白名单详情。")
def get_admin_access_rule(
    ruleId: Annotated[str, Path(description="黑白名单规则 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    return ok(access_rule_item(get_access_rule_or_404(db, ruleId)))


@router.put("/access-rules/{ruleId}", response_model=AdminResponse, summary="修改黑白名单", description="修改接口限流黑名单或白名单规则。")
def update_admin_access_rule(
    ruleId: Annotated[str, Path(description="黑白名单规则 ID")],
    payload: AdminAccessRulePayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    rule = get_access_rule_or_404(db, ruleId)
    rule.rule_type = payload.type
    rule.ip = (payload.ip or "").strip() or None
    rule.method = (payload.method or "ALL").upper()
    rule.path = (payload.path or "").strip() or None
    rule.status = payload.status
    rule.remark = payload.remark or ""
    rule.last_time = utc_now()
    db.commit()
    db.refresh(rule)
    return ok(access_rule_item(rule), "黑白名单规则更新成功")


@router.delete("/access-rules/{ruleId}", response_model=AdminResponse, summary="删除黑白名单", description="删除接口限流黑名单或白名单规则。")
def delete_admin_access_rule(
    ruleId: Annotated[str, Path(description="黑白名单规则 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    rule = get_access_rule_or_404(db, ruleId)
    db.delete(rule)
    db.commit()
    return ok(None, "黑白名单规则删除成功")


@router.get(
    "/account/profile",
    response_model=AdminResponse,
    summary="当前后台账号资料",
    description="个人中心 - 个人信息接口，返回当前登录后台账号最新资料。",
)
def get_current_admin_profile(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    return ok(account_item(get_current_admin_account(db, admin_subject)))


@router.put(
    "/account/profile",
    response_model=AdminResponse,
    summary="修改当前后台账号资料",
    description="个人中心 - 修改昵称、头像、邮箱、手机号和备注；用户名、角色和权限不在这里修改。",
)
def update_current_admin_profile(
    payload: AdminAccountPayload,
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
    request: Request,
) -> dict[str, Any]:
    account = get_current_admin_account(db, admin_subject)
    if payload.nickname is not None:
        account.nickname = payload.nickname
    if payload.avatar is not None:
        account.avatar = payload.avatar
    if payload.email is not None:
        account.email = payload.email
    if payload.phone is not None:
        account.phone = payload.phone
    if payload.remark is not None:
        account.remark = payload.remark
    account.last_time = utc_now()
    write_admin_log(
        db,
        account=account,
        username=account.username,
        category="account",
        action="profile_update",
        title="个人信息更新",
        content="修改当前后台账号个人资料",
        status_value="success",
        request=request,
    )
    db.commit()
    db.refresh(account)
    return ok(account_item(account), "个人信息更新成功")


@router.get(
    "/logs",
    response_model=AdminResponse,
    summary="日志管理列表",
    description="日志管理列表接口，支持按关键词、日志分类、状态和账号筛选。",
)
def list_admin_logs(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="标题、内容、动作、IP、归属地关键词")] = None,
    category: Annotated[str | None, Query(description="日志分类：login/security/account 等")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="日志状态：success/warning/error")] = None,
    username: Annotated[str | None, Query(description="后台账号用户名")] = None,
    accountId: Annotated[str | None, Query(description="后台账号 ID")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=200, description="每页数量")] = 10,
) -> dict[str, Any]:
    query = db.query(AdminOperationLog)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter(
            or_(
                AdminOperationLog.title.ilike(like_keyword),
                AdminOperationLog.content.ilike(like_keyword),
                AdminOperationLog.action.ilike(like_keyword),
                AdminOperationLog.ip.ilike(like_keyword),
                AdminOperationLog.location.ilike(like_keyword),
            )
        )
    if category:
        query = query.filter(AdminOperationLog.category == category)
    if status_filter:
        query = query.filter(AdminOperationLog.status == status_filter)
    if username:
        query = query.filter(AdminOperationLog.username.ilike(f"%{username}%"))
    if accountId:
        query = query.filter(AdminOperationLog.account_id == accountId)

    total = query.count()
    logs = query.order_by(AdminOperationLog.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [operation_log_item(log) for log in logs], "total": total})


@router.get(
    "/logs/{logId}",
    response_model=AdminResponse,
    summary="日志详情",
    description="根据日志 ID 查询日志详情。",
)
def get_admin_log(
    logId: Annotated[str, Path(description="日志 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    return ok(operation_log_item(get_log_or_404(db, logId)))


@router.delete(
    "/logs/{logId}",
    response_model=AdminResponse,
    summary="删除日志",
    description="删除单条后台操作日志。",
)
def delete_admin_log(
    logId: Annotated[str, Path(description="日志 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    log = get_log_or_404(db, logId)
    db.delete(log)
    db.commit()
    return ok(None, "日志删除成功")


@router.post(
    "/logs/batch-delete",
    response_model=AdminResponse,
    summary="批量删除日志",
    description="按日志 ID 批量删除后台操作日志。",
)
def batch_delete_admin_logs(
    payload: BatchIdsPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    count = db.query(AdminOperationLog).filter(AdminOperationLog.log_id.in_(payload.ids)).delete(synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量删除成功")


@router.get(
    "/notifications",
    response_model=AdminResponse,
    summary="后台通知下拉",
    description="PC 后台右上角消息通知下拉接口，返回未读数量和最近通知列表。",
)
def list_admin_notifications(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
    limit: Annotated[int, Query(ge=1, le=50, description="返回数量")] = 10,
) -> dict[str, Any]:
    admin_user_id = f"admin:{admin_subject}"
    recent_system_messages = db.query(AdminSystemMessage).order_by(AdminSystemMessage.create_time.desc()).limit(limit).all()
    for system_message in recent_system_messages:
        ensure_admin_notification(db, admin_subject, system_message)
    db.commit()

    query = db.query(Message).filter(Message.user_id == admin_user_id)
    unread_count = query.filter(Message.is_read.is_(False)).count()
    messages = query.order_by(Message.create_time.desc()).limit(limit).all()
    return ok({"unreadCount": unread_count, "list": [notification_item(message) for message in messages]})


@router.get(
    "/notifications/{messageId}",
    response_model=AdminResponse,
    summary="后台通知详情",
    description="PC 后台点击右上角通知时调用；按当前 admin 标记该条通知为已读并返回详情。",
)
def get_admin_notification_detail(
    messageId: Annotated[str, Path(description="通知消息 ID")],
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    admin_user_id = f"admin:{admin_subject}"
    message = db.query(Message).filter(Message.user_id == admin_user_id, Message.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "通知未找到")
    message.is_read = True
    message.last_time = utc_now()
    db.commit()
    db.refresh(message)
    return ok(notification_item(message), "已读取通知详情")


@router.put(
    "/notifications/read-all",
    response_model=AdminResponse,
    summary="后台通知全部已读",
    description="PC 后台右上角消息通知下拉全部标记已读。",
)
def read_all_admin_notifications(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    admin_user_id = f"admin:{admin_subject}"
    db.query(Message).filter(Message.user_id == admin_user_id, Message.is_read.is_(False)).update(
        {Message.is_read: True, Message.last_time: utc_now()},
        synchronize_session=False,
    )
    db.commit()
    return ok(None, "已全部标记为已读")


@router.put(
    "/notifications/{messageId}/read",
    response_model=AdminResponse,
    summary="后台通知已读",
    description="PC 后台右上角消息通知下拉单条标记已读。",
)
def read_admin_notification(
    messageId: Annotated[str, Path(description="通知消息 ID")],
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    admin_user_id = f"admin:{admin_subject}"
    message = db.query(Message).filter(Message.user_id == admin_user_id, Message.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "通知未找到")
    message.is_read = True
    message.last_time = utc_now()
    db.commit()
    return ok(notification_item(message), "已标记为已读")


@router.get(
    "/users",
    response_model=AdminResponse,
    summary="用户列表",
    description="后台用户管理列表，支持按用户 ID、昵称、手机号、账号状态筛选。",
)
def list_admin_users(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    userId: Annotated[str | None, Query(description="业务用户 ID，精确匹配")] = None,
    user_id_legacy: Annotated[str | None, Query(alias="user_id", description="兼容旧参数 user_id", include_in_schema=False)] = None,
    nickname: Annotated[str | None, Query(description="用户昵称，模糊匹配")] = None,
    phone: Annotated[str | None, Query(description="手机号，模糊匹配")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="账号状态：normal 正常，banned 禁用")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    user_id = userId or user_id_legacy
    query = db.query(User)
    if user_id:
        query = query.filter(User.user_id == user_id)
    if nickname:
        query = query.filter(User.nickname.ilike(f"%{nickname}%"))
    if phone:
        phone_value = normalize_phone(phone) or phone
        query = query.filter(or_(User.phone.ilike(f"%{phone_value}%"), User.user_id.ilike(f"%{phone_value}%")))
    if status_filter == "normal":
        query = query.filter(User.is_active.is_(True))
    elif status_filter == "banned":
        query = query.filter(User.is_active.is_(False))

    total = query.count()
    users = query.order_by(User.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [user_item(db, user) for user in users], "total": total})


@router.get(
    "/user-online",
    response_model=AdminResponse,
    summary="用户端在线状态",
    description="系统配置 - 在线用户列表。status 默认 all，可传 online/offline/all 查看在线、离线或全部用户端用户。",
)
def list_admin_user_online(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    status_filter: Annotated[str, Query(alias="status", pattern="^(all|online|offline)$", description="在线状态：all 全部，online 在线，offline 离线")] = "all",
    keyword: Annotated[str | None, Query(description="用户 ID、昵称、手机号关键词")] = None,
    userId: Annotated[str | None, Query(description="业务用户 ID，精确匹配")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    now = utc_now()
    cutoff = now - timedelta(seconds=USER_ONLINE_WINDOW_SECONDS)
    base_query = db.query(User).filter(User.is_active.is_(True))
    if userId:
        base_query = base_query.filter(User.user_id == userId)
    if keyword:
        normalized_keyword = normalize_phone(keyword) or keyword
        base_query = base_query.filter(
            or_(
                User.user_id.ilike(f"%{keyword}%"),
                User.nickname.ilike(f"%{keyword}%"),
                User.phone.ilike(f"%{normalized_keyword}%"),
                User.new_phone.ilike(f"%{normalized_keyword}%"),
            )
        )

    online_condition = User.last_time >= cutoff
    online_count = base_query.filter(online_condition).count()
    total_all = base_query.count()
    offline_count = total_all - online_count

    query = base_query
    if status_filter == "online":
        query = query.filter(online_condition)
    elif status_filter == "offline":
        query = query.filter(~online_condition)

    total = query.count()
    users = query.order_by(User.last_time.desc(), User.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok(
        {
            "list": [user_online_item(user, now) for user in users],
            "total": total,
            "onlineCount": online_count,
            "offlineCount": offline_count,
            "status": status_filter,
            "onlineWindowSeconds": USER_ONLINE_WINDOW_SECONDS,
            "page": page,
            "pageSize": limit,
        }
    )


@router.get(
    "/users/export",
    summary="导出用户",
    description="导出用户列表，支持 .xls 和 .xlsx；默认导出 .xls。",
)
def export_admin_users(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    format: Annotated[str, Query(pattern="^(xls|xlsx)$", description="导出格式：xls 或 xlsx")] = "xls",
    nickname: Annotated[str | None, Query(description="用户昵称，模糊匹配")] = None,
    phone: Annotated[str | None, Query(description="手机号，模糊匹配")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="账号状态：normal 正常，banned 禁用")] = None,
) -> StreamingResponse:
    query = db.query(User)
    if nickname:
        query = query.filter(User.nickname.ilike(f"%{nickname}%"))
    if phone:
        phone_value = normalize_phone(phone) or phone
        query = query.filter(or_(User.phone.ilike(f"%{phone_value}%"), User.user_id.ilike(f"%{phone_value}%")))
    if status_filter == "normal":
        query = query.filter(User.is_active.is_(True))
    elif status_filter == "banned":
        query = query.filter(User.is_active.is_(False))

    users = query.order_by(User.create_time.desc()).all()
    return user_export_response(users, format)


@router.post(
    "/users/import",
    response_model=AdminResponse,
    summary="导入用户",
    description="导入用户 Excel，支持 .xls 和 .xlsx。表头支持中文或驼峰字段，例如 用户ID、昵称、手机号、账号状态。",
)
def import_admin_users(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    file: Annotated[UploadFile, File(description="用户 Excel 文件，仅支持 .xls/.xlsx")],
) -> dict[str, Any]:
    rows = read_user_import_rows(file.filename or "", file.file.read())
    return ok(import_users_from_rows(db, rows), "用户导入成功")


@router.get(
    "/export",
    summary="后台通用导出",
    description="后台通用导出入口。target/type 默认 users，当前支持导出用户 .xls/.xlsx。",
)
def export_admin_data(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    target: Annotated[str, Query(description="导出类型，默认 users")] = "users",
    type_alias: Annotated[str | None, Query(alias="type", description="兼容旧参数 type", include_in_schema=False)] = None,
    format: Annotated[str, Query(pattern="^(xls|xlsx)$", description="导出格式：xls 或 xlsx")] = "xls",
    nickname: Annotated[str | None, Query(description="用户昵称，模糊匹配")] = None,
    phone: Annotated[str | None, Query(description="手机号，模糊匹配")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="账号状态：normal 正常，banned 禁用")] = None,
) -> StreamingResponse:
    export_target = (type_alias or target or "users").strip().lower()
    if export_target not in {"users", "user"}:
        raise fail(status.HTTP_400_BAD_REQUEST, "暂不支持该导出类型")

    query = db.query(User)
    if nickname:
        query = query.filter(User.nickname.ilike(f"%{nickname}%"))
    if phone:
        query = query.filter(User.phone.ilike(f"%{phone}%"))
    if status_filter == "normal":
        query = query.filter(User.is_active.is_(True))
    elif status_filter == "banned":
        query = query.filter(User.is_active.is_(False))
    users = query.order_by(User.create_time.desc()).all()
    return user_export_response(users, format)


@router.post(
    "/import",
    response_model=AdminResponse,
    summary="后台通用导入",
    description="后台通用导入入口。target/type 默认 users，当前支持导入用户 .xls/.xlsx。",
)
def import_admin_data(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    file: Annotated[UploadFile, File(description="Excel 文件，仅支持 .xls/.xlsx")],
    target: Annotated[str, Query(description="导入类型，默认 users")] = "users",
    type_alias: Annotated[str | None, Query(alias="type", description="兼容旧参数 type", include_in_schema=False)] = None,
) -> dict[str, Any]:
    import_target = (type_alias or target or "users").strip().lower()
    if import_target not in {"users", "user"}:
        raise fail(status.HTTP_400_BAD_REQUEST, "暂不支持该导入类型")
    rows = read_user_import_rows(file.filename or "", file.file.read())
    return ok(import_users_from_rows(db, rows), "用户导入成功")


@router.get(
    "/users/{userId}",
    response_model=AdminResponse,
    summary="用户详情",
    description="后台查看单个业务用户详情。",
)
def get_admin_user(
    userId: Annotated[str, Path(description="业务用户 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    user_id = userId
    user = db.query(User).filter(User.user_id == user_id).one_or_none()
    if user is None:
        raise fail(status.HTTP_404_NOT_FOUND, "用户未找到")
    return ok(user_item(db, user))


@router.put(
    "/users/{userId}",
    response_model=AdminResponse,
    summary="修改用户状态",
    description="后台禁用或解禁用户。禁用后该用户 token 将无法访问需要登录的用户端接口。",
)
def update_admin_user_status(
    userId: Annotated[str, Path(description="业务用户 ID")],
    payload: AdminUserStatusUpdate,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    user_id = userId
    user = db.query(User).filter(User.user_id == user_id).one_or_none()
    if user is None:
        raise fail(status.HTTP_404_NOT_FOUND, "用户未找到")
    user.is_active = payload.status == "normal"
    user.last_time = utc_now() if user.is_active else utc_now() - timedelta(seconds=USER_ONLINE_WINDOW_SECONDS + 1)
    db.commit()
    return ok(None, "操作成功")


@router.get(
    "/posts",
    response_model=AdminResponse,
    summary="内容列表",
    description="后台内容管理列表，支持按发布人、内容关键词和上架状态筛选。",
)
def list_admin_posts(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    nickname: Annotated[str | None, Query(description="发布人昵称，模糊匹配")] = None,
    userId: Annotated[str | None, Query(description="发布人业务用户 ID，精确匹配")] = None,
    user_id_legacy: Annotated[str | None, Query(alias="user_id", description="兼容旧参数 user_id", include_in_schema=False)] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="内容状态：online 已上架，offline 已下架，draft 草稿")] = None,
    content: Annotated[str | None, Query(description="内容正文关键词，模糊匹配")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 5,
) -> dict[str, Any]:
    user_id = userId or user_id_legacy
    query = db.query(Post).options(joinedload(Post.author)).filter(Post.is_deleted.is_(False))
    if user_id:
        query = query.filter(Post.user_id == user_id)
    if status_filter == "online":
        query = query.filter(Post.status.in_(("online", "published")))
    elif status_filter == "offline":
        query = query.filter(Post.status.in_(("offline", "draft")))
    elif status_filter == "draft":
        query = query.filter(Post.status == "draft")
    if content:
        query = query.filter(Post.content.ilike(f"%{content}%"))
    if nickname:
        query = query.join(User, User.user_id == Post.user_id).filter(User.nickname.ilike(f"%{nickname}%"))

    total = query.count()
    posts = query.order_by(Post.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [post_item(post) for post in posts], "total": total})


@router.get(
    "/posts/{postId}",
    response_model=AdminResponse,
    summary="内容详情",
    description="后台查看单条内容详情；后台可查看已下架内容。",
)
def get_admin_post(
    postId: Annotated[str, Path(description="内容 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    post_id = postId
    post = (
        db.query(Post)
        .options(joinedload(Post.author))
        .filter(Post.post_id == post_id, Post.is_deleted.is_(False))
        .one_or_none()
    )
    if post is None:
        raise fail(status.HTTP_404_NOT_FOUND, "内容未找到")
    return ok(post_item(post))


@router.put(
    "/posts/{postId}/offline",
    response_model=AdminResponse,
    summary="内容下架",
    description="后台将内容置为 offline。下架后用户端公开列表和详情不可见。",
)
def offline_admin_post(
    postId: Annotated[str, Path(description="内容 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    post_id = postId
    post = db.query(Post).filter(Post.post_id == post_id, Post.is_deleted.is_(False)).one_or_none()
    if post is None:
        raise fail(status.HTTP_404_NOT_FOUND, "内容未找到")
    post.status = "offline"
    post.last_time = utc_now()
    db.commit()
    return ok(None, "内容已成功下架")


@router.put(
    "/posts/{postId}/restore",
    response_model=AdminResponse,
    summary="恢复上架",
    description="后台将已下架内容恢复为 online，恢复后用户端可见。",
)
def restore_admin_post(
    postId: Annotated[str, Path(description="内容 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    post_id = postId
    post = db.query(Post).filter(Post.post_id == post_id, Post.is_deleted.is_(False)).one_or_none()
    if post is None:
        raise fail(status.HTTP_404_NOT_FOUND, "内容未找到")
    post.status = "online"
    post.last_time = utc_now()
    db.commit()
    return ok(None, "内容已恢复上架")


@router.get(
    "/comments",
    response_model=AdminResponse,
    summary="评论列表",
    description="后台评论管理列表，支持按内容 ID 或评论人用户 ID 筛选。",
)
def list_admin_comments(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    postId: Annotated[str | None, Query(description="内容 ID，精确匹配")] = None,
    userId: Annotated[str | None, Query(description="评论人业务用户 ID，精确匹配")] = None,
    post_id_legacy: Annotated[str | None, Query(alias="post_id", description="兼容旧参数 post_id", include_in_schema=False)] = None,
    user_id_legacy: Annotated[str | None, Query(alias="user_id", description="兼容旧参数 user_id", include_in_schema=False)] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    post_id = postId or post_id_legacy
    user_id = userId or user_id_legacy
    query = db.query(Comment).filter(Comment.is_deleted.is_(False))
    if post_id:
        query = query.filter(Comment.post_id == post_id)
    if user_id:
        query = query.filter(Comment.user_id == user_id)
    total = query.count()
    comments = query.order_by(Comment.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [comment_item(db, comment) for comment in comments], "total": total})


@router.delete(
    "/comments/{commentId}",
    response_model=AdminResponse,
    summary="删除评论",
    description="后台删除评论。当前为软删除，并同步扣减内容评论数。",
)
def delete_admin_comment(
    commentId: Annotated[str, Path(description="评论 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    comment_id = commentId
    comment = db.query(Comment).filter(Comment.comment_id == comment_id, Comment.is_deleted.is_(False)).one_or_none()
    if comment is None:
        raise fail(status.HTTP_404_NOT_FOUND, "评论未找到")
    post = db.query(Post).filter(Post.post_id == comment.post_id).one_or_none()
    comment.is_deleted = True
    comment.delete_time = utc_now()
    comment.last_time = comment.delete_time
    if post is not None:
        post.comment_count = max(0, post.comment_count - 1)
        post.last_time = utc_now()
    db.commit()
    return ok(None, "评论已成功删除")


@router.get(
    "/likes",
    response_model=AdminResponse,
    summary="点赞列表",
    description="后台点赞管理列表，支持按内容 ID、点赞人用户 ID、内容或昵称关键词筛选。",
)
def list_admin_likes(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    postId: Annotated[str | None, Query(description="内容 ID，精确匹配")] = None,
    userId: Annotated[str | None, Query(description="点赞人业务用户 ID，精确匹配")] = None,
    post_id_legacy: Annotated[str | None, Query(alias="post_id", description="兼容旧参数 post_id", include_in_schema=False)] = None,
    user_id_legacy: Annotated[str | None, Query(alias="user_id", description="兼容旧参数 user_id", include_in_schema=False)] = None,
    keyword: Annotated[str | None, Query(description="内容、昵称、用户 ID 关键词")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    post_id = postId or post_id_legacy
    user_id = userId or user_id_legacy
    query = db.query(PostLike)
    if post_id:
        query = query.filter(PostLike.post_id == post_id)
    if user_id:
        query = query.filter(PostLike.user_id == user_id)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.join(Post, Post.post_id == PostLike.post_id).join(User, User.user_id == PostLike.user_id).filter(
            or_(Post.content.ilike(like_keyword), User.nickname.ilike(like_keyword), User.user_id.ilike(like_keyword))
        )
    total = query.count()
    likes = query.order_by(PostLike.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [like_item(db, like) for like in likes], "total": total})


@router.delete(
    "/likes/{likeId}",
    response_model=AdminResponse,
    summary="删除点赞记录",
    description="后台删除点赞记录，并同步扣减内容点赞数。",
)
def delete_admin_like(
    likeId: Annotated[int, Path(description="点赞记录 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    like = db.query(PostLike).filter(PostLike.id == likeId).one_or_none()
    if like is None:
        raise fail(status.HTTP_404_NOT_FOUND, "点赞记录未找到")
    post = db.query(Post).filter(Post.post_id == like.post_id).one_or_none()
    if post is not None:
        post.like_count = max(0, post.like_count - 1)
        post.last_time = utc_now()
    db.delete(like)
    db.commit()
    return ok(None, "点赞记录已删除")


@router.get(
    "/shares",
    response_model=AdminResponse,
    summary="分享列表",
    description="后台分享管理列表，支持按内容 ID、分享人用户 ID、平台、场景、内容或昵称关键词筛选。",
)
def list_admin_shares(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    postId: Annotated[str | None, Query(description="内容 ID，精确匹配")] = None,
    userId: Annotated[str | None, Query(description="分享人业务用户 ID，精确匹配")] = None,
    post_id_legacy: Annotated[str | None, Query(alias="post_id", description="兼容旧参数 post_id", include_in_schema=False)] = None,
    user_id_legacy: Annotated[str | None, Query(alias="user_id", description="兼容旧参数 user_id", include_in_schema=False)] = None,
    platform: Annotated[str | None, Query(description="分享平台，例如 h5/wechat")] = None,
    scene: Annotated[str | None, Query(description="分享场景")] = None,
    keyword: Annotated[str | None, Query(description="内容、昵称、用户 ID 关键词")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    post_id = postId or post_id_legacy
    user_id = userId or user_id_legacy
    query = db.query(PostShare)
    if post_id:
        query = query.filter(PostShare.post_id == post_id)
    if user_id:
        query = query.filter(PostShare.user_id == user_id)
    if platform:
        query = query.filter(PostShare.platform.ilike(f"%{platform}%"))
    if scene:
        query = query.filter(PostShare.scene.ilike(f"%{scene}%"))
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.join(Post, Post.post_id == PostShare.post_id).join(User, User.user_id == PostShare.user_id).filter(
            or_(Post.content.ilike(like_keyword), User.nickname.ilike(like_keyword), User.user_id.ilike(like_keyword))
        )
    total = query.count()
    shares = query.order_by(PostShare.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [share_item(db, share) for share in shares], "total": total})


@router.delete(
    "/shares/{shareId}",
    response_model=AdminResponse,
    summary="删除分享记录",
    description="后台删除分享记录，并同步扣减内容分享数。",
)
def delete_admin_share(
    shareId: Annotated[int, Path(description="分享记录 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    share = db.query(PostShare).filter(PostShare.id == shareId).one_or_none()
    if share is None:
        raise fail(status.HTTP_404_NOT_FOUND, "分享记录未找到")
    post = db.query(Post).filter(Post.post_id == share.post_id).one_or_none()
    if post is not None:
        post.share_count = max(0, post.share_count - 1)
        post.last_time = utc_now()
    db.delete(share)
    db.commit()
    return ok(None, "分享记录已删除")


@router.get(
    "/agreement/{agreementType}",
    response_model=AdminResponse,
    summary="获取协议",
    description="获取后台维护的协议内容。agreementType 支持 privacy 或 user。",
)
def get_admin_agreement(
    agreementType: Annotated[str, Path(description="协议类型：privacy 隐私协议，user 用户协议")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    agreement_type = agreementType
    if agreement_type not in DEFAULT_AGREEMENTS:
        raise fail(status.HTTP_404_NOT_FOUND, "协议未找到")
    agreement = get_or_create_agreement(db, agreement_type)
    return ok(agreement.content)


@router.put(
    "/agreement/{agreementType}",
    response_model=AdminResponse,
    summary="更新协议",
    description="更新后台维护的协议内容。agreementType 支持 privacy 或 user。",
)
def update_admin_agreement(
    agreementType: Annotated[str, Path(description="协议类型：privacy 隐私协议，user 用户协议")],
    payload: AgreementUpdate,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    agreement_type = agreementType
    if agreement_type not in DEFAULT_AGREEMENTS:
        raise fail(status.HTTP_404_NOT_FOUND, "协议未找到")
    agreement = get_or_create_agreement(db, agreement_type)
    agreement.content = payload.content
    agreement.last_time = utc_now()
    db.commit()
    return ok(None, "协议更新成功")


def seed_announcements_if_empty(db: Session) -> None:
    if db.query(AdminAnnouncement).count():
        return
    db.add_all(
        [
            AdminAnnouncement(
                announcement_id=new_business_id("ann"),
                title="欢迎使用即闪后台",
                type="info",
                content="<p>即闪平台后台公告模块已启用。</p>",
                link="",
                pinned=True,
                start_time=format_time(utc_now()),
                end_time="",
                status="active",
            ),
            AdminAnnouncement(
                announcement_id=new_business_id("ann"),
                title="内容安全巡检提醒",
                type="warning",
                content="<p>请定期关注内容审核、评论举报和标签风险数据。</p>",
                link="",
                pinned=False,
                start_time=format_time(utc_now()),
                end_time="",
                status="active",
            ),
        ]
    )
    db.commit()


def seed_versions_if_empty(db: Session) -> None:
    now = format_time(utc_now())
    defaults = {
        "iOS": ("1.0.0", "100"),
        "Android": ("1.0.0", "100"),
        "HarmonyOS": ("1.0.0", "100"),
    }
    existing_platforms = {row[0] for row in db.query(AdminVersion.platform).filter(AdminVersion.status != "deprecated").distinct().all()}
    missing_versions = [
        AdminVersion(
            version_id=new_business_id("ver"),
            platform=platform,
            version=version,
            build=build,
            force_update=False,
            status="released",
            beta_pct=100,
            notes="首个稳定版本",
            notes_type="text",
            download_url="",
            release_time=now,
        )
        for platform, (version, build) in defaults.items()
        if platform not in existing_platforms
    ]
    if missing_versions:
        db.add_all(missing_versions)
        db.commit()


def seed_accounts_if_empty(db: Session) -> None:
    if db.query(AdminAccount).count():
        return
    db.add(
        AdminAccount(
            account_id=new_business_id("acc"),
            username="admin",
            nickname="超级管理员",
            avatar="",
            role="superadmin",
            permissions=DEFAULT_ADMIN_PERMISSIONS,
            status="active",
            email="admin@example.com",
            phone="13800000000",
            remark="系统默认管理员",
            password="123456",
            last_login=format_time(utc_now()),
        )
    )
    db.commit()


def seed_permissions_if_empty(db: Session) -> None:
    existing_ids = {row[0] for row in db.query(AdminPermission.permission_id).all()}
    existing_keys = {row[0] for row in db.query(AdminPermission.permission_key).all()}
    new_permissions = []
    for item in DEFAULT_ADMIN_PERMISSION_MODULES:
        if item["permission_id"] in existing_ids or item["permission_key"] in existing_keys:
            continue
        new_permissions.append(
            AdminPermission(
                permission_id=item["permission_id"],
                permission_key=item["permission_key"],
                label=item["label"],
                description=item.get("description"),
                sort=item.get("sort", 0),
                status=item.get("status", "enabled"),
                is_default=item.get("is_default", False),
                remark=item.get("remark"),
            )
        )
    if new_permissions:
        db.add_all(new_permissions)
        db.commit()


def seed_roles_if_empty(db: Session) -> None:
    existing_ids = {row[0] for row in db.query(AdminRole.role_id).all()}
    existing_keys = {row[0] for row in db.query(AdminRole.role_key).all()}
    new_roles = []
    for item in DEFAULT_ADMIN_ROLES:
        if item["role_id"] in existing_ids or item["role_key"] in existing_keys:
            continue
        new_roles.append(
            AdminRole(
                role_id=item["role_id"],
                role_key=item["role_key"],
                label=item["label"],
                icon=item.get("icon"),
                permissions=item.get("permissions", []),
                sort=item.get("sort", 0),
                status=item.get("status", "enabled"),
                is_default=item.get("is_default", False),
                remark=item.get("remark"),
            )
        )
    if new_roles:
        db.add_all(new_roles)
        db.commit()


def seed_feedback_form_if_empty(db: Session) -> None:
    config = db.query(AdminFeedbackFormConfig).filter(AdminFeedbackFormConfig.config_id == DEFAULT_FEEDBACK_CONFIG_ID).one_or_none()
    if config is None:
        config = AdminFeedbackFormConfig(
            config_id=DEFAULT_FEEDBACK_CONFIG_ID,
            title="意见反馈",
            menu_title="反馈管理",
            description="请填写手机号码、标题和反馈内容，我们会尽快处理。",
            submit_button_text="提交反馈",
            success_message="反馈提交成功",
            status="enabled",
            remark="系统默认反馈表单",
        )
        db.add(config)
        db.commit()
    existing_keys = {
        row[0]
        for row in db.query(AdminFeedbackField.field_key)
        .filter(AdminFeedbackField.form_id == DEFAULT_FEEDBACK_CONFIG_ID)
        .all()
    }
    new_fields = [
        AdminFeedbackField(
            field_id=item["field_id"],
            form_id=DEFAULT_FEEDBACK_CONFIG_ID,
            field_key=item["field_key"],
            label=item["label"],
            type=item["type"],
            placeholder=item["placeholder"],
            required=item["required"],
            options=[],
            sort=item["sort"],
            status="enabled",
            is_default=True,
            remark="系统默认反馈字段",
        )
        for item in DEFAULT_FEEDBACK_FIELDS
        if item["field_key"] not in existing_keys
    ]
    if new_fields:
        db.add_all(new_fields)
        db.commit()


def active_permission_keys(db: Session) -> set[str]:
    seed_permissions_if_empty(db)
    return {row[0] for row in db.query(AdminPermission.permission_key).filter(AdminPermission.status == "enabled").all()}


def account_effective_permissions(db: Session, account: AdminAccount) -> list[str]:
    active_keys = active_permission_keys(db)
    if account.role == "superadmin":
        return [item["permission_key"] for item in DEFAULT_ADMIN_PERMISSION_MODULES if item["permission_key"] in active_keys]
    return [key for key in account.permissions or [] if key in active_keys]


def validate_permission_keys(db: Session, permissions: list[str] | None) -> None:
    if not permissions:
        return
    enabled_keys = active_permission_keys(db)
    missing_keys = [key for key in permissions if key not in enabled_keys]
    if missing_keys:
        raise fail(status.HTTP_400_BAD_REQUEST, f"权限模块不存在或已停用：{', '.join(missing_keys)}")


def seed_menus_if_empty(db: Session) -> None:
    existing_ids = {row[0] for row in db.query(AdminMenu.menu_id).all()}
    new_menus = []
    for item in DEFAULT_ADMIN_MENUS:
        if item["menu_id"] in existing_ids:
            continue
        new_menus.append(
            AdminMenu(
                menu_id=item["menu_id"],
                parent_id=item.get("parent_id"),
                title=item["title"],
                path=item["path"],
                name=item["name"],
                component=item.get("component"),
                redirect=item.get("redirect"),
                icon=item.get("icon"),
                type=item.get("type", "menu"),
                permission=item.get("permission"),
                sort=item.get("sort", 0),
                status=item.get("status", "enabled"),
                visible=item.get("visible", True),
                keep_alive=item.get("keep_alive", False),
                affix=item.get("affix", False),
                external_link=item.get("external_link"),
                remark=item.get("remark"),
            )
        )
    if new_menus:
        db.add_all(new_menus)
        db.commit()


def get_announcement_or_404(db: Session, announcement_id: str) -> AdminAnnouncement:
    announcement = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id == announcement_id).one_or_none()
    if announcement is None:
        raise fail(status.HTTP_404_NOT_FOUND, "公告未找到")
    return announcement


def get_or_create_single_banner(db: Session) -> AdminAnnouncement:
    announcement = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id == SINGLE_BANNER_ANNOUNCEMENT_ID).one_or_none()
    if announcement is not None:
        return announcement
    announcement = AdminAnnouncement(
        announcement_id=SINGLE_BANNER_ANNOUNCEMENT_ID,
        title="【系统通知】即闪 App 服务升级公告",
        type="info",
        content="<p>请在后台维护 App 首页单公告内容。</p>",
        link="",
        pinned=True,
        start_time=format_time(utc_now()),
        end_time="",
        status="active",
    )
    db.add(announcement)
    db.commit()
    db.refresh(announcement)
    return announcement


def get_version_or_404(db: Session, version_id: str) -> AdminVersion:
    version = db.query(AdminVersion).filter(AdminVersion.version_id == version_id).one_or_none()
    if version is None:
        raise fail(status.HTTP_404_NOT_FOUND, "版本未找到")
    return version


def get_package_or_404(db: Session, package_id: str) -> AdminPackage:
    package = db.query(AdminPackage).filter(AdminPackage.package_id == package_id).one_or_none()
    if package is None:
        raise fail(status.HTTP_404_NOT_FOUND, "安装包未找到")
    return package


def safe_upload_filename(filename: str) -> str:
    name = FilePath(filename or "package.bin").name
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name) or "package.bin"


def save_package_upload(file: UploadFile, package_id: str, platform: str) -> tuple[str, str, str, int, str]:
    original_filename = safe_upload_filename(file.filename or "package.bin")
    suffix = FilePath(original_filename).suffix
    upload_dir = PACKAGE_UPLOAD_ROOT / platform.lower()
    upload_dir.mkdir(parents=True, exist_ok=True)
    relative_path = FilePath("uploads") / "packages" / platform.lower() / f"{package_id}{suffix}"
    target_path = PACKAGE_UPLOAD_ROOT.parent.parent / relative_path

    md5 = hashlib.md5()
    size_bytes = 0
    with target_path.open("wb") as output:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            size_bytes += len(chunk)
            md5.update(chunk)
            output.write(chunk)
    if size_bytes == 0:
        target_path.unlink(missing_ok=True)
        raise fail(status.HTTP_400_BAD_REQUEST, "上传文件不能为空")
    file_path = f"static/{relative_path.as_posix()}"
    download_url = f"/static/{relative_path.as_posix()}"
    return original_filename, file_path, download_url, size_bytes, md5.hexdigest()


def get_account_or_404(db: Session, account_id: str) -> AdminAccount:
    account = db.query(AdminAccount).filter(AdminAccount.account_id == account_id).one_or_none()
    if account is None:
        raise fail(status.HTTP_404_NOT_FOUND, "账号未找到")
    return account


def get_role_or_404(db: Session, role_id: str) -> AdminRole:
    role = db.query(AdminRole).filter(AdminRole.role_id == role_id).one_or_none()
    if role is None:
        raise fail(status.HTTP_404_NOT_FOUND, "角色未找到")
    return role


def get_permission_or_404(db: Session, permission_id: str) -> AdminPermission:
    permission = db.query(AdminPermission).filter(AdminPermission.permission_id == permission_id).one_or_none()
    if permission is None:
        raise fail(status.HTTP_404_NOT_FOUND, "权限模块未找到")
    return permission


def get_log_or_404(db: Session, log_id: str) -> AdminOperationLog:
    log = db.query(AdminOperationLog).filter(AdminOperationLog.log_id == log_id).one_or_none()
    if log is None:
        raise fail(status.HTTP_404_NOT_FOUND, "日志未找到")
    return log


def permission_key_in_use(db: Session, permission_key: str) -> bool:
    account_permissions = [row[0] or [] for row in db.query(AdminAccount.permissions).all()]
    role_permissions = [row[0] or [] for row in db.query(AdminRole.permissions).all()]
    return any(permission_key in keys for keys in [*account_permissions, *role_permissions])


def validate_role_exists(db: Session, role_key: str | None) -> None:
    if not role_key:
        return
    seed_roles_if_empty(db)
    role = db.query(AdminRole).filter(AdminRole.role_key == role_key).one_or_none()
    if role is None:
        raise fail(status.HTTP_400_BAD_REQUEST, "角色不存在")
    if role.status != "enabled":
        raise fail(status.HTTP_400_BAD_REQUEST, "角色已停用")


def get_menu_or_404(db: Session, menu_id: str) -> AdminMenu:
    menu = db.query(AdminMenu).filter(AdminMenu.menu_id == menu_id).one_or_none()
    if menu is None:
        raise fail(status.HTTP_404_NOT_FOUND, "菜单未找到")
    return menu


def get_current_admin_account(db: Session, username: str) -> AdminAccount:
    seed_accounts_if_empty(db)
    account = db.query(AdminAccount).filter(AdminAccount.username == username).one_or_none()
    if account is None:
        raise fail(status.HTTP_401_UNAUTHORIZED, "登录账号不存在，请重新登录")
    if account.status != "active":
        raise fail(status.HTTP_403_FORBIDDEN, "账号已禁用，请联系管理员")
    return account


def get_or_create_security_setting(db: Session, account: AdminAccount) -> AdminSecuritySetting:
    setting = db.query(AdminSecuritySetting).filter(AdminSecuritySetting.account_id == account.account_id).one_or_none()
    if setting is not None:
        return setting
    setting = AdminSecuritySetting(account_id=account.account_id, mfa_enabled=False, password_policy_enabled=True, remark="")
    db.add(setting)
    db.commit()
    db.refresh(setting)
    return setting


def get_access_rule_or_404(db: Session, rule_id: str) -> AdminAccessRule:
    rule = db.query(AdminAccessRule).filter(AdminAccessRule.rule_id == rule_id).one_or_none()
    if rule is None:
        raise fail(status.HTTP_404_NOT_FOUND, "黑白名单规则未找到")
    return rule


def request_ip(request: Request | None) -> str:
    if request is None or request.client is None:
        return "127.0.0.1"
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host or "127.0.0.1"


def request_user_agent(request: Request | None) -> str:
    if request is None:
        return ""
    return request.headers.get("user-agent", "")


def ip_location(ip: str) -> str:
    if ip.startswith("127.") or ip == "::1" or ip.startswith("192.168.") or ip.startswith("10."):
        return "中国·浙江·杭州"
    return "未知"


def write_admin_log(
    db: Session,
    *,
    account: AdminAccount | None,
    username: str | None,
    category: str,
    action: str,
    title: str,
    content: str = "",
    status_value: str = "success",
    request: Request | None = None,
) -> AdminOperationLog:
    ip = request_ip(request)
    log = AdminOperationLog(
        log_id=new_business_id("log"),
        account_id=account.account_id if account is not None else None,
        username=account.username if account is not None else username,
        category=category,
        action=action,
        title=title,
        content=content,
        status=status_value,
        ip=ip,
        location=ip_location(ip),
        user_agent=request_user_agent(request),
    )
    db.add(log)
    return log


def validate_menu_parent(db: Session, parent_id: str | None, self_id: str | None = None) -> None:
    if not parent_id:
        return
    if parent_id == self_id:
        raise fail(status.HTTP_400_BAD_REQUEST, "上级菜单不能选择自己")
    parent = get_menu_or_404(db, parent_id)
    cursor = parent
    while cursor.parent_id:
        if cursor.parent_id == self_id:
            raise fail(status.HTTP_400_BAD_REQUEST, "上级菜单不能选择自己的下级")
        cursor = get_menu_or_404(db, cursor.parent_id)


def validate_menu_unique(db: Session, name: str, path: str, self_id: str | None = None) -> None:
    name_query = db.query(AdminMenu).filter(AdminMenu.name == name)
    path_query = db.query(AdminMenu).filter(AdminMenu.path == path)
    if self_id:
        name_query = name_query.filter(AdminMenu.menu_id != self_id)
        path_query = path_query.filter(AdminMenu.menu_id != self_id)
    if name_query.one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "路由名称已存在")
    if path_query.one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "路由路径已存在")


@router.get("/announcements", response_model=AdminResponse, summary="公告列表", description="公告管理列表，支持关键词、类型、状态筛选。")
def list_admin_announcements(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="标题或内容关键词")] = None,
    type: Annotated[str | None, Query(description="公告类型：info/warning/danger")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：active/inactive")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=10000, description="每页数量")] = 10,
) -> dict[str, Any]:
    seed_announcements_if_empty(db)
    query = db.query(AdminAnnouncement)
    if keyword:
        query = query.filter((AdminAnnouncement.title.ilike(f"%{keyword}%")) | (AdminAnnouncement.content.ilike(f"%{keyword}%")))
    if type:
        query = query.filter(AdminAnnouncement.type == type)
    if status_filter:
        query = query.filter(AdminAnnouncement.status == status_filter)
    total = query.count()
    items = query.order_by(AdminAnnouncement.pinned.desc(), AdminAnnouncement.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [announcement_item(item) for item in items], "total": total})


@router.post("/announcements", response_model=AdminResponse, summary="新增公告", description="新增后台公告。")
def create_admin_announcement(payload: AdminAnnouncementPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    if not payload.title or not payload.content:
        raise fail(status.HTTP_400_BAD_REQUEST, "公告标题和内容不能为空")
    announcement = AdminAnnouncement(
        announcement_id=new_business_id("ann"),
        title=payload.title,
        type=payload.type or "info",
        content=payload.content,
        link=payload.link or "",
        pinned=bool(payload.pinned),
        start_time=payload.startTime or format_time(utc_now()),
        end_time=payload.endTime or "",
        status=payload.status or "active",
    )
    db.add(announcement)
    db.commit()
    db.refresh(announcement)
    return ok(announcement_item(announcement), "公告创建成功")


@router.put("/announcements/batch-publish", response_model=AdminResponse, summary="批量发布公告", description="批量设置公告为发布状态。")
def batch_publish_announcements_first(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id.in_(payload.ids)).update({"status": "active", "last_time": utc_now()}, synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量发布成功")


@router.put("/announcements/batch-disable", response_model=AdminResponse, summary="批量停用公告", description="批量设置公告为停用状态。")
def batch_disable_announcements_first(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id.in_(payload.ids)).update({"status": "inactive", "last_time": utc_now()}, synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量停用成功")


@router.post("/announcements/batch-delete", response_model=AdminResponse, summary="批量删除公告", description="批量删除后台公告。")
def batch_delete_announcements_first(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id.in_(payload.ids)).delete(synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量删除成功")


@router.get("/announcements/{announcementId}", response_model=AdminResponse, summary="公告详情", description="根据公告 ID 查询公告详情。SINGLE_BANNER 为 App 首页单公告。")
def get_admin_announcement(announcementId: Annotated[str, Path(description="公告 ID；SINGLE_BANNER 表示 App 首页单公告")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    if announcementId == SINGLE_BANNER_ANNOUNCEMENT_ID:
        announcement = get_or_create_single_banner(db)
    else:
        announcement = get_announcement_or_404(db, announcementId)
    return ok(announcement_item(announcement))


@router.put("/announcements/{announcementId}", response_model=AdminResponse, summary="修改公告", description="修改后台公告，支持局部字段更新。")
def update_admin_announcement(announcementId: Annotated[str, Path(description="公告 ID")], payload: AdminAnnouncementPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    if announcementId == SINGLE_BANNER_ANNOUNCEMENT_ID:
        announcement = get_or_create_single_banner(db)
    else:
        announcement = get_announcement_or_404(db, announcementId)
    if payload.title is not None:
        announcement.title = payload.title
    if payload.type is not None:
        announcement.type = payload.type
    if payload.content is not None:
        announcement.content = payload.content
    if payload.link is not None:
        announcement.link = payload.link
    if payload.pinned is not None:
        announcement.pinned = payload.pinned
    if payload.startTime is not None:
        announcement.start_time = payload.startTime
    if payload.endTime is not None:
        announcement.end_time = payload.endTime
    if payload.status is not None:
        announcement.status = payload.status
    announcement.last_time = utc_now()
    db.commit()
    db.refresh(announcement)
    return ok(announcement_item(announcement), "公告更新成功")


@router.delete("/announcements/{announcementId}", response_model=AdminResponse, summary="删除公告", description="删除后台公告。")
def delete_admin_announcement(announcementId: Annotated[str, Path(description="公告 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    announcement = get_announcement_or_404(db, announcementId)
    db.delete(announcement)
    db.commit()
    return ok(None, "公告删除成功")


@router.put("/announcements/batch-publish", response_model=AdminResponse, summary="批量发布公告", description="批量设置公告为发布状态。")
def batch_publish_announcements(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id.in_(payload.ids)).update({"status": "active", "last_time": utc_now()}, synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量发布成功")


@router.put("/announcements/batch-disable", response_model=AdminResponse, summary="批量停用公告", description="批量设置公告为停用状态。")
def batch_disable_announcements(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id.in_(payload.ids)).update({"status": "inactive", "last_time": utc_now()}, synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量停用成功")


@router.post("/announcements/batch-delete", response_model=AdminResponse, summary="批量删除公告", description="批量删除后台公告。")
def batch_delete_announcements(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminAnnouncement).filter(AdminAnnouncement.announcement_id.in_(payload.ids)).delete(synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量删除成功")


@router.get("/versions", response_model=AdminResponse, summary="版本列表", description="版本管理列表，支持平台、状态、强制更新筛选。")
def list_admin_versions(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    platform: Annotated[str | None, Query(description="平台：iOS/Android/HarmonyOS")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：released/beta/deprecated")] = None,
    forceUpdate: Annotated[bool | None, Query(description="是否强制更新")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=10000, description="每页数量")] = 10,
) -> dict[str, Any]:
    seed_versions_if_empty(db)
    query = db.query(AdminVersion)
    if platform:
        query = query.filter(AdminVersion.platform == platform)
    if status_filter:
        query = query.filter(AdminVersion.status == status_filter)
    if forceUpdate is not None:
        query = query.filter(AdminVersion.force_update == forceUpdate)
    total = query.count()
    items = query.order_by(AdminVersion.platform.asc(), AdminVersion.build.desc(), AdminVersion.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [version_item(item) for item in items], "total": total})


@router.post("/versions", response_model=AdminResponse, summary="新增版本", description="新增 App 版本记录。")
def create_admin_version(payload: AdminVersionPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    if not payload.platform or not payload.version or not payload.build:
        raise fail(status.HTTP_400_BAD_REQUEST, "平台、版本号和 Build 号不能为空")
    version = AdminVersion(
        version_id=new_business_id("ver"),
        platform=payload.platform,
        version=payload.version,
        build=payload.build,
        force_update=bool(payload.forceUpdate),
        status=payload.status or "released",
        beta_pct=payload.betaPct if payload.betaPct is not None else 0,
        notes=payload.notes or "",
        notes_type=payload.notesType or "text",
        download_url=payload.downloadUrl or "",
        release_time=format_time(utc_now()),
    )
    db.add(version)
    db.commit()
    db.refresh(version)
    return ok(version_item(version), "版本创建成功")


@router.put("/versions/batch-deprecate", response_model=AdminResponse, summary="批量下线版本", description="批量设置 App 版本为已下线。")
def batch_deprecate_versions_first(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminVersion).filter(AdminVersion.version_id.in_(payload.ids)).update({"status": "deprecated", "last_time": utc_now()}, synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量下线成功")


@router.post("/versions/batch-delete", response_model=AdminResponse, summary="批量删除版本", description="批量删除 App 版本记录。")
def batch_delete_versions_first(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminVersion).filter(AdminVersion.version_id.in_(payload.ids)).delete(synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量删除成功")


@router.put("/versions/{versionId}", response_model=AdminResponse, summary="修改版本", description="修改 App 版本记录，支持局部字段更新。")
def update_admin_version(versionId: Annotated[str, Path(description="版本 ID")], payload: AdminVersionPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    version = get_version_or_404(db, versionId)
    if payload.platform is not None:
        version.platform = payload.platform
    if payload.version is not None:
        version.version = payload.version
    if payload.build is not None:
        version.build = payload.build
    if payload.forceUpdate is not None:
        version.force_update = payload.forceUpdate
    if payload.status is not None:
        version.status = payload.status
    if payload.betaPct is not None:
        version.beta_pct = payload.betaPct
    if payload.notes is not None:
        version.notes = payload.notes
    if payload.notesType is not None:
        version.notes_type = payload.notesType
    if payload.downloadUrl is not None:
        version.download_url = payload.downloadUrl
    version.last_time = utc_now()
    db.commit()
    db.refresh(version)
    return ok(version_item(version), "版本更新成功")


@router.delete("/versions/{versionId}", response_model=AdminResponse, summary="删除版本", description="删除 App 版本记录。")
def delete_admin_version(versionId: Annotated[str, Path(description="版本 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    version = get_version_or_404(db, versionId)
    db.delete(version)
    db.commit()
    return ok(None, "版本删除成功")


@router.put("/versions/{versionId}/deprecate", response_model=AdminResponse, summary="下线版本", description="将 App 版本状态设置为已下线。")
def deprecate_admin_version(versionId: Annotated[str, Path(description="版本 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    version = get_version_or_404(db, versionId)
    version.status = "deprecated"
    version.last_time = utc_now()
    db.commit()
    db.refresh(version)
    return ok(version_item(version), "版本已下线")


@router.put("/versions/batch-deprecate", response_model=AdminResponse, summary="批量下线版本", description="批量设置 App 版本为已下线。")
def batch_deprecate_versions(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminVersion).filter(AdminVersion.version_id.in_(payload.ids)).update({"status": "deprecated", "last_time": utc_now()}, synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量下线成功")


@router.post("/versions/batch-delete", response_model=AdminResponse, summary="批量删除版本", description="批量删除 App 版本记录。")
def batch_delete_versions(payload: BatchIdsPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    count = db.query(AdminVersion).filter(AdminVersion.version_id.in_(payload.ids)).delete(synchronize_session=False)
    db.commit()
    return ok({"count": count}, "批量删除成功")


@router.get("/packages", response_model=AdminResponse, summary="安装包列表", description="安装包管理列表，默认展示 Android，可按平台、版本号和关键词筛选。")
def list_admin_packages(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    platform: Annotated[str, Query(pattern="^(Android|iOS|HarmonyOS)$", description="平台，默认 Android")] = "Android",
    version: Annotated[str | None, Query(description="版本号，精确匹配")] = None,
    keyword: Annotated[str | None, Query(description="包名、原始文件名、MD5 关键词")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：uploaded/active/deprecated")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=1000, description="每页数量")] = 10,
) -> dict[str, Any]:
    query = db.query(AdminPackage).filter(AdminPackage.platform == platform)
    if version:
        query = query.filter(AdminPackage.version == version)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter(
            or_(
                AdminPackage.display_name.ilike(like_keyword),
                AdminPackage.original_filename.ilike(like_keyword),
                AdminPackage.md5.ilike(like_keyword),
            )
        )
    if status_filter:
        query = query.filter(AdminPackage.status == status_filter)

    total = query.count()
    items = query.order_by(AdminPackage.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    latest = db.query(AdminPackage).filter(AdminPackage.platform == platform).order_by(AdminPackage.create_time.desc()).first()
    version_rows = (
        db.query(AdminPackage.version, func.count(AdminPackage.id), func.coalesce(func.sum(AdminPackage.size_bytes), 0), func.max(AdminPackage.create_time))
        .filter(AdminPackage.platform == platform)
        .group_by(AdminPackage.version)
        .order_by(func.max(AdminPackage.create_time).desc())
        .all()
    )
    return ok(
        {
            "platform": platform,
            "list": [package_item(item) for item in items],
            "total": total,
            "latest": package_item(latest) if latest else None,
            "versions": [package_version_summary_item(row) for row in version_rows],
            "platformTabs": ["Android", "iOS", "HarmonyOS"],
        }
    )


@router.post("/packages/upload", response_model=AdminResponse, summary="上传安装包", description="上传 Android/iOS/HarmonyOS 安装包，后端自动计算 MD5 和文件大小。")
def upload_admin_package(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
    request: Request,
    platform: Annotated[str, Form(pattern="^(Android|iOS|HarmonyOS)$", description="平台：Android/iOS/HarmonyOS")],
    version: Annotated[str, Form(min_length=1, max_length=32, description="版本号")],
    file: Annotated[UploadFile, File(description="安装包文件")],
    build: Annotated[str | None, Form(max_length=32, description="Build 号")] = None,
    displayName: Annotated[str | None, Form(max_length=256, description="安装包展示名称")] = None,
    remark: Annotated[str | None, Form(description="备注")] = None,
) -> dict[str, Any]:
    account = get_current_admin_account(db, admin_subject)
    package_id = new_business_id("pkg")
    original_filename, file_path, download_url, size_bytes, md5 = save_package_upload(file, package_id, platform)
    package = AdminPackage(
        package_id=package_id,
        platform=platform,
        version=version,
        build=build or "",
        display_name=displayName or original_filename,
        original_filename=original_filename,
        file_path=file_path,
        download_url=download_url,
        md5=md5,
        size_bytes=size_bytes,
        status="uploaded",
        remark=remark or "",
    )
    db.add(package)
    write_admin_log(
        db,
        account=account,
        username=account.username,
        category="package",
        action="package_upload",
        title="上传安装包",
        content=f"{platform} {version} {original_filename}",
        status_value="success",
        request=request,
    )
    db.commit()
    db.refresh(package)
    return ok(package_item(package), "安装包上传成功")


@router.get("/packages/history", response_model=AdminResponse, summary="版本安装包历史", description="点击某个版本时查询该版本下所有安装包上传记录。")
def list_admin_package_history(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    platform: Annotated[str, Query(pattern="^(Android|iOS|HarmonyOS)$", description="平台")] = "Android",
    version: Annotated[str, Query(min_length=1, description="版本号")] = "",
) -> dict[str, Any]:
    query = db.query(AdminPackage).filter(AdminPackage.platform == platform, AdminPackage.version == version)
    items = query.order_by(AdminPackage.create_time.desc()).all()
    return ok({"platform": platform, "version": version, "list": [package_item(item) for item in items], "total": len(items)})


@router.get("/packages/{packageId}", response_model=AdminResponse, summary="安装包详情", description="根据安装包 ID 查询文件名、MD5、大小和版本等详情。")
def get_admin_package(
    packageId: Annotated[str, Path(description="安装包 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    return ok(package_item(get_package_or_404(db, packageId)))


@router.put("/packages/{packageId}", response_model=AdminResponse, summary="修改安装包", description="修改安装包展示名称、版本号、Build 号、状态和备注。")
def update_admin_package(
    packageId: Annotated[str, Path(description="安装包 ID")],
    payload: AdminPackageUpdatePayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    package = get_package_or_404(db, packageId)
    if payload.displayName is not None:
        package.display_name = payload.displayName
    if payload.version is not None:
        package.version = payload.version
    if payload.build is not None:
        package.build = payload.build
    if payload.status is not None:
        package.status = payload.status
    if payload.remark is not None:
        package.remark = payload.remark
    package.last_time = utc_now()
    db.commit()
    db.refresh(package)
    return ok(package_item(package), "安装包更新成功")


@router.delete("/packages/{packageId}", response_model=AdminResponse, summary="删除安装包", description="删除安装包记录。文件会保留在 static 目录，避免下载链接被误删。")
def delete_admin_package(
    packageId: Annotated[str, Path(description="安装包 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    package = get_package_or_404(db, packageId)
    db.delete(package)
    db.commit()
    return ok(None, "安装包删除成功")


@router.get("/menus", response_model=AdminResponse, summary="菜单列表", description="后台动态菜单列表，默认返回树形结构。")
def list_admin_menus(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="菜单标题、路径或路由名称关键词")] = None,
    parentId: Annotated[str | None, Query(description="上级菜单 ID；传入后只查询直接下级")] = None,
    type: Annotated[str | None, Query(description="菜单类型：catalog/menu/button/link")] = None,
    permission: Annotated[str | None, Query(description="权限标识")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled/disabled")] = None,
    visible: Annotated[bool | None, Query(description="是否在菜单显示")] = None,
    tree: Annotated[bool, Query(description="是否返回树形 children")] = True,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=1000, description="每页数量")] = 1000,
) -> dict[str, Any]:
    seed_menus_if_empty(db)
    query = db.query(AdminMenu)
    if keyword:
        query = query.filter((AdminMenu.title.ilike(f"%{keyword}%")) | (AdminMenu.path.ilike(f"%{keyword}%")) | (AdminMenu.name.ilike(f"%{keyword}%")))
    if parentId is not None:
        query = query.filter(AdminMenu.parent_id == parentId)
    if type:
        query = query.filter(AdminMenu.type == type)
    if permission:
        query = query.filter(AdminMenu.permission == permission)
    if status_filter:
        query = query.filter(AdminMenu.status == status_filter)
    if visible is not None:
        query = query.filter(AdminMenu.visible == visible)
    total = query.count()
    menus = query.order_by(AdminMenu.sort.asc(), AdminMenu.create_time.asc()).offset((page - 1) * limit).limit(limit).all()
    if tree and parentId is None:
        return ok({"list": menu_tree_items(menus), "total": total})
    return ok({"list": [menu_item(menu) for menu in menus], "total": total})


@router.get("/menus/tree", response_model=AdminResponse, summary="当前账号菜单树", description="根据当前登录账号权限返回可显示的动态菜单树。")
def current_admin_menu_tree(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_menus_if_empty(db)
    active_keys = active_permission_keys(db)
    account = get_current_admin_account(db, admin_subject)
    menus = db.query(AdminMenu).order_by(AdminMenu.sort.asc(), AdminMenu.create_time.asc()).all()
    tree = permitted_menu_tree_items(menus, account, active_keys)
    return ok({"list": tree, "total": len(flatten_menu_tree(tree)), "permissions": account_effective_permissions(db, account), "role": account.role})


@router.get("/menus/routes", response_model=AdminResponse, summary="动态路由", description="根据当前登录账号权限返回前端动态路由和菜单树。")
def current_admin_menu_routes(
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_menus_if_empty(db)
    active_keys = active_permission_keys(db)
    account = get_current_admin_account(db, admin_subject)
    menus = db.query(AdminMenu).order_by(AdminMenu.sort.asc(), AdminMenu.create_time.asc()).all()
    tree = permitted_menu_tree_items(menus, account, active_keys)
    flat_list = flatten_menu_tree(tree)
    return ok(
        {
            "list": tree,
            "routes": tree,
            "flatList": flat_list,
            "permissions": account_effective_permissions(db, account),
            "role": account.role,
            "username": account.username,
        }
    )


@router.post("/menus", response_model=AdminResponse, summary="新增菜单", description="新增后台动态菜单或动态路由配置。")
def create_admin_menu(payload: AdminMenuPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_menus_if_empty(db)
    validate_permission_keys(db, [payload.permission] if payload.permission else None)
    validate_menu_parent(db, payload.parentId)
    validate_menu_unique(db, payload.name, payload.path)
    menu = AdminMenu(
        menu_id=new_business_id("menu"),
        parent_id=payload.parentId,
        title=payload.title,
        path=payload.path,
        name=payload.name,
        component=payload.component,
        redirect=payload.redirect,
        icon=payload.icon,
        type=payload.type,
        permission=payload.permission,
        sort=payload.sort,
        status=payload.status,
        visible=payload.visible,
        keep_alive=payload.keepAlive,
        affix=payload.affix,
        external_link=payload.externalLink,
        remark=payload.remark,
    )
    db.add(menu)
    db.commit()
    db.refresh(menu)
    return ok(menu_item(menu), "菜单创建成功")


@router.get("/menus/{menuId}", response_model=AdminResponse, summary="菜单详情", description="根据菜单 ID 查询菜单详情。")
def get_admin_menu(menuId: Annotated[str, Path(description="业务菜单 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_menus_if_empty(db)
    return ok(menu_item(get_menu_or_404(db, menuId)))


@router.put("/menus/{menuId}", response_model=AdminResponse, summary="修改菜单", description="修改后台动态菜单或动态路由配置。")
def update_admin_menu(menuId: Annotated[str, Path(description="业务菜单 ID")], payload: AdminMenuPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_menus_if_empty(db)
    validate_permission_keys(db, [payload.permission] if payload.permission else None)
    menu = get_menu_or_404(db, menuId)
    validate_menu_parent(db, payload.parentId, self_id=menuId)
    validate_menu_unique(db, payload.name, payload.path, self_id=menuId)
    menu.parent_id = payload.parentId
    menu.title = payload.title
    menu.path = payload.path
    menu.name = payload.name
    menu.component = payload.component
    menu.redirect = payload.redirect
    menu.icon = payload.icon
    menu.type = payload.type
    menu.permission = payload.permission
    menu.sort = payload.sort
    menu.status = payload.status
    menu.visible = payload.visible
    menu.keep_alive = payload.keepAlive
    menu.affix = payload.affix
    menu.external_link = payload.externalLink
    menu.remark = payload.remark
    menu.last_time = utc_now()
    db.commit()
    db.refresh(menu)
    return ok(menu_item(menu), "菜单更新成功")


@router.delete("/menus/{menuId}", response_model=AdminResponse, summary="删除菜单", description="删除动态菜单；存在下级时不允许删除。")
def delete_admin_menu(menuId: Annotated[str, Path(description="业务菜单 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_menus_if_empty(db)
    menu = get_menu_or_404(db, menuId)
    if db.query(AdminMenu).filter(AdminMenu.parent_id == menuId).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "存在下级菜单，不能删除")
    db.delete(menu)
    db.commit()
    return ok(None, "菜单删除成功")


@router.get("/permissions", response_model=AdminResponse, summary="权限模块列表", description="账号管理 - 权限模块列表；账号勾选后才会展示对应菜单。")
def list_admin_permissions(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="权限名称或 key 关键词")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled/disabled")] = None,
    onlyActive: Annotated[bool, Query(description="是否只返回启用权限，账号新增/编辑弹窗建议传 true")] = False,
) -> dict[str, Any]:
    seed_permissions_if_empty(db)
    query = db.query(AdminPermission)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter((AdminPermission.label.ilike(like_keyword)) | (AdminPermission.permission_key.ilike(like_keyword)))
    if status_filter:
        query = query.filter(AdminPermission.status == status_filter)
    if onlyActive:
        query = query.filter(AdminPermission.status == "enabled")
    permissions = query.order_by(AdminPermission.sort.asc(), AdminPermission.create_time.asc()).all()
    return ok({"list": [permission_item(permission) for permission in permissions], "total": len(permissions)})


@router.post("/permissions", response_model=AdminResponse, summary="新增权限模块", description="新增账号可勾选的权限模块。")
def create_admin_permission(payload: AdminPermissionPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_permissions_if_empty(db)
    if db.query(AdminPermission).filter(AdminPermission.permission_key == payload.permissionKey).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "权限标识已存在")
    permission = AdminPermission(
        permission_id=new_business_id("perm"),
        permission_key=payload.permissionKey,
        label=payload.label,
        description=payload.description,
        sort=payload.sort,
        status=payload.status,
        is_default=False,
        remark=payload.remark,
    )
    db.add(permission)
    db.commit()
    db.refresh(permission)
    return ok(permission_item(permission), "权限模块创建成功")


@router.get("/permissions/{permissionId}", response_model=AdminResponse, summary="权限模块详情", description="根据权限模块 ID 查询详情。")
def get_admin_permission(permissionId: Annotated[str, Path(description="权限模块 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_permissions_if_empty(db)
    return ok(permission_item(get_permission_or_404(db, permissionId)))


@router.put("/permissions/{permissionId}", response_model=AdminResponse, summary="修改权限模块", description="修改权限模块。系统默认权限允许改名称/状态，但不允许改标识。")
def update_admin_permission(permissionId: Annotated[str, Path(description="权限模块 ID")], payload: AdminPermissionPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_permissions_if_empty(db)
    permission = get_permission_or_404(db, permissionId)
    if permission.is_default and payload.permissionKey != permission.permission_key:
        raise fail(status.HTTP_400_BAD_REQUEST, "系统默认权限标识不能修改")
    if payload.permissionKey != permission.permission_key and permission_key_in_use(db, permission.permission_key):
        raise fail(status.HTTP_400_BAD_REQUEST, "权限正在使用中，不能修改标识")
    duplicate = db.query(AdminPermission).filter(AdminPermission.permission_key == payload.permissionKey, AdminPermission.permission_id != permissionId).one_or_none()
    if duplicate is not None:
        raise fail(status.HTTP_400_BAD_REQUEST, "权限标识已存在")
    permission.permission_key = payload.permissionKey
    permission.label = payload.label
    permission.description = payload.description
    permission.sort = payload.sort
    permission.status = payload.status
    permission.remark = payload.remark
    permission.last_time = utc_now()
    db.commit()
    db.refresh(permission)
    return ok(permission_item(permission), "权限模块更新成功")


@router.delete("/permissions/{permissionId}", response_model=AdminResponse, summary="删除权限模块", description="删除自定义权限模块；系统默认权限不允许删除。")
def delete_admin_permission(permissionId: Annotated[str, Path(description="权限模块 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_permissions_if_empty(db)
    permission = get_permission_or_404(db, permissionId)
    if permission.is_default:
        raise fail(status.HTTP_400_BAD_REQUEST, "系统默认权限不能删除")
    if permission_key_in_use(db, permission.permission_key):
        raise fail(status.HTTP_400_BAD_REQUEST, "存在角色或账号正在使用该权限，不能删除")
    db.delete(permission)
    db.commit()
    return ok(None, "权限模块删除成功")


@router.get("/roles", response_model=AdminResponse, summary="后台角色列表", description="账号管理 - 角色列表，默认包含超级管理员、管理员、运营员、观察员。")
def list_admin_roles(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="角色名称或 key 关键词")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled/disabled")] = None,
) -> dict[str, Any]:
    seed_roles_if_empty(db)
    query = db.query(AdminRole)
    if keyword:
        query = query.filter((AdminRole.label.ilike(f"%{keyword}%")) | (AdminRole.role_key.ilike(f"%{keyword}%")))
    if status_filter:
        query = query.filter(AdminRole.status == status_filter)
    roles = query.order_by(AdminRole.is_default.desc(), AdminRole.sort.asc(), AdminRole.create_time.asc()).all()
    return ok({"list": [role_item(role) for role in roles], "total": len(roles)})


@router.post("/roles", response_model=AdminResponse, summary="新增后台角色", description="新增后台角色，默认角色不可删除，自定义角色可删除。")
def create_admin_role(payload: AdminRolePayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_roles_if_empty(db)
    validate_permission_keys(db, payload.permissions)
    if db.query(AdminRole).filter(AdminRole.role_key == payload.roleKey).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "角色标识已存在")
    custom_role_count = db.query(AdminRole).filter(AdminRole.is_default.is_(False)).count()
    if custom_role_count >= 3:
        raise fail(status.HTTP_400_BAD_REQUEST, "自定义角色最多只能新增 3 个")
    role = AdminRole(
        role_id=new_business_id("role"),
        role_key=payload.roleKey,
        label=payload.label,
        icon=payload.icon,
        permissions=payload.permissions,
        sort=payload.sort,
        status=payload.status,
        is_default=False,
        remark=payload.remark,
    )
    db.add(role)
    db.commit()
    db.refresh(role)
    return ok(role_item(role), "角色创建成功")


@router.get("/roles/{roleId}", response_model=AdminResponse, summary="后台角色详情", description="根据角色 ID 查询角色详情。")
def get_admin_role(roleId: Annotated[str, Path(description="角色 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_roles_if_empty(db)
    return ok(role_item(get_role_or_404(db, roleId)))


@router.put("/roles/{roleId}", response_model=AdminResponse, summary="修改后台角色", description="修改后台角色，默认角色也允许修改，但不允许删除。")
def update_admin_role(roleId: Annotated[str, Path(description="角色 ID")], payload: AdminRolePayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_roles_if_empty(db)
    validate_permission_keys(db, payload.permissions)
    role = get_role_or_404(db, roleId)
    duplicate = db.query(AdminRole).filter(AdminRole.role_key == payload.roleKey, AdminRole.role_id != roleId).one_or_none()
    if duplicate is not None:
        raise fail(status.HTTP_400_BAD_REQUEST, "角色标识已存在")
    role.role_key = payload.roleKey
    role.label = payload.label
    role.icon = payload.icon
    role.permissions = payload.permissions
    role.sort = payload.sort
    role.status = payload.status
    role.remark = payload.remark
    role.last_time = utc_now()
    db.commit()
    db.refresh(role)
    return ok(role_item(role), "角色更新成功")


@router.delete("/roles/{roleId}", response_model=AdminResponse, summary="删除后台角色", description="删除自定义角色；系统默认角色不允许删除。")
def delete_admin_role(roleId: Annotated[str, Path(description="角色 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_roles_if_empty(db)
    role = get_role_or_404(db, roleId)
    if role.is_default:
        raise fail(status.HTTP_400_BAD_REQUEST, "系统默认角色不能删除")
    if db.query(AdminAccount).filter(AdminAccount.role == role.role_key).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "存在账号正在使用该角色，不能删除")
    db.delete(role)
    db.commit()
    return ok(None, "角色删除成功")


@router.get("/accounts", response_model=AdminResponse, summary="后台账号列表", description="账号管理列表。")
def list_admin_accounts(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="搜索用户名、昵称、邮箱、手机号")] = None,
    role: Annotated[str | None, Query(description="角色 key")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：active/disabled")] = None,
) -> dict[str, Any]:
    seed_accounts_if_empty(db)
    query = db.query(AdminAccount)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter(
            (AdminAccount.username.ilike(like_keyword))
            | (AdminAccount.nickname.ilike(like_keyword))
            | (AdminAccount.email.ilike(like_keyword))
            | (AdminAccount.phone.ilike(like_keyword))
        )
    if role:
        query = query.filter(AdminAccount.role == role)
    if status_filter:
        query = query.filter(AdminAccount.status == status_filter)
    accounts = query.order_by(AdminAccount.role.asc(), AdminAccount.create_time.desc()).all()
    return ok([account_item(account) for account in accounts])


@router.post("/accounts", response_model=AdminResponse, summary="新增后台账号", description="新增后台管理账号。")
def create_admin_account(payload: AdminAccountPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    if not payload.username or not payload.nickname:
        raise fail(status.HTTP_400_BAD_REQUEST, "账号和昵称不能为空")
    if db.query(AdminAccount).filter(AdminAccount.username == payload.username).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "账号已存在")
    validate_role_exists(db, payload.role or "admin")
    validate_permission_keys(db, payload.permissions)
    account = AdminAccount(
        account_id=new_business_id("acc"),
        username=payload.username,
        nickname=payload.nickname,
        avatar=payload.avatar or "",
        role=payload.role or "admin",
        permissions=payload.permissions or [],
        status=payload.status or "active",
        email=payload.email or "",
        phone=payload.phone or "",
        remark=payload.remark or "",
        password="123456",
        last_login="",
    )
    db.add(account)
    db.commit()
    db.refresh(account)
    return ok(account_item(account), "账号创建成功")


@router.get("/accounts/{accountId}", response_model=AdminResponse, summary="后台账号详情", description="根据账号 ID 查询后台账号最新详情。")
def get_admin_account(accountId: Annotated[str, Path(description="账号 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    seed_accounts_if_empty(db)
    return ok(account_item(get_account_or_404(db, accountId)))


@router.put("/accounts/{accountId}", response_model=AdminResponse, summary="修改后台账号", description="修改后台账号信息。")
def update_admin_account(accountId: Annotated[str, Path(description="账号 ID")], payload: AdminAccountPayload, db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    account = get_account_or_404(db, accountId)
    if payload.username is not None and payload.username != account.username:
        if db.query(AdminAccount).filter(AdminAccount.username == payload.username).one_or_none():
            raise fail(status.HTTP_400_BAD_REQUEST, "账号已存在")
        account.username = payload.username
    if payload.nickname is not None:
        account.nickname = payload.nickname
    if payload.avatar is not None:
        account.avatar = payload.avatar
    if payload.role is not None:
        validate_role_exists(db, payload.role)
        account.role = payload.role
    if payload.permissions is not None:
        validate_permission_keys(db, payload.permissions)
        account.permissions = payload.permissions
    if payload.status is not None:
        account.status = payload.status
    if payload.email is not None:
        account.email = payload.email
    if payload.phone is not None:
        account.phone = payload.phone
    if payload.remark is not None:
        account.remark = payload.remark
    account.last_time = utc_now()
    db.commit()
    db.refresh(account)
    return ok(account_item(account), "账号更新成功")


@router.delete("/accounts/{accountId}", response_model=AdminResponse, summary="删除后台账号", description="删除后台账号。")
def delete_admin_account(accountId: Annotated[str, Path(description="账号 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    account = get_account_or_404(db, accountId)
    db.delete(account)
    db.commit()
    return ok(None, "账号删除成功")


@router.post("/accounts/{accountId}/reset-password", response_model=AdminResponse, summary="重置后台账号密码", description="将后台账号密码重置为 123456。")
def reset_admin_account_password(accountId: Annotated[str, Path(description="账号 ID")], db: Annotated[Session, Depends(get_db)], _: Annotated[str, Depends(get_admin_subject)]) -> dict[str, Any]:
    account = get_account_or_404(db, accountId)
    account.password = "123456"
    account.last_time = utc_now()
    db.commit()
    return ok(None, "密码已重置为 123456")


@router.get("/feedback/config", response_model=AdminResponse, summary="反馈表单配置", description="系统配置 - 反馈管理动态表单配置。")
def get_admin_feedback_config(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    config = db.query(AdminFeedbackFormConfig).filter(AdminFeedbackFormConfig.config_id == DEFAULT_FEEDBACK_CONFIG_ID).one()
    return ok(feedback_config_item(db, config))


@router.put("/feedback/config", response_model=AdminResponse, summary="修改反馈表单配置", description="修改用户端反馈动态表单页面标题、说明和按钮文案。")
def update_admin_feedback_config(
    payload: AdminFeedbackConfigPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    config = db.query(AdminFeedbackFormConfig).filter(AdminFeedbackFormConfig.config_id == DEFAULT_FEEDBACK_CONFIG_ID).one()
    config.title = payload.title
    config.menu_title = payload.menuTitle
    config.description = payload.description
    config.submit_button_text = payload.submitButtonText
    config.success_message = payload.successMessage
    config.status = payload.status
    config.remark = payload.remark
    config.last_time = utc_now()
    db.commit()
    db.refresh(config)
    return ok(feedback_config_item(db, config), "反馈表单配置已更新")


@router.get("/feedback/fields", response_model=AdminResponse, summary="反馈字段列表", description="系统配置 - 反馈动态表单字段列表。")
def list_admin_feedback_fields(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled/disabled")] = None,
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    query = db.query(AdminFeedbackField).filter(AdminFeedbackField.form_id == DEFAULT_FEEDBACK_CONFIG_ID)
    if status_filter:
        query = query.filter(AdminFeedbackField.status == status_filter)
    fields = query.order_by(AdminFeedbackField.sort.asc(), AdminFeedbackField.create_time.asc()).all()
    return ok({"list": [feedback_field_item(field) for field in fields], "total": len(fields)})


@router.post("/feedback/fields", response_model=AdminResponse, summary="新增反馈字段", description="新增用户端反馈动态表单字段。")
def create_admin_feedback_field(
    payload: AdminFeedbackFieldPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    duplicate = (
        db.query(AdminFeedbackField)
        .filter(AdminFeedbackField.form_id == DEFAULT_FEEDBACK_CONFIG_ID, AdminFeedbackField.field_key == payload.fieldKey)
        .one_or_none()
    )
    if duplicate is not None:
        raise fail(status.HTTP_400_BAD_REQUEST, "反馈字段标识已存在")
    field = AdminFeedbackField(
        field_id=new_business_id("fbf"),
        form_id=DEFAULT_FEEDBACK_CONFIG_ID,
        field_key=payload.fieldKey,
        label=payload.label,
        type=payload.type,
        placeholder=payload.placeholder,
        required=payload.required,
        options=payload.options,
        sort=payload.sort,
        status=payload.status,
        is_default=False,
        remark=payload.remark,
    )
    db.add(field)
    db.commit()
    db.refresh(field)
    return ok(feedback_field_item(field), "反馈字段创建成功")


@router.get("/feedback/fields/{fieldId}", response_model=AdminResponse, summary="反馈字段详情", description="根据字段 ID 查询反馈动态字段。")
def get_admin_feedback_field(
    fieldId: Annotated[str, Path(description="反馈字段 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    field = db.query(AdminFeedbackField).filter(AdminFeedbackField.field_id == fieldId).one_or_none()
    if field is None:
        raise fail(status.HTTP_404_NOT_FOUND, "反馈字段未找到")
    return ok(feedback_field_item(field))


@router.put("/feedback/fields/{fieldId}", response_model=AdminResponse, summary="修改反馈字段", description="修改反馈动态表单字段。默认字段允许改名称、必填、排序和状态，但不允许改字段标识。")
def update_admin_feedback_field(
    fieldId: Annotated[str, Path(description="反馈字段 ID")],
    payload: AdminFeedbackFieldPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    field = db.query(AdminFeedbackField).filter(AdminFeedbackField.field_id == fieldId).one_or_none()
    if field is None:
        raise fail(status.HTTP_404_NOT_FOUND, "反馈字段未找到")
    if field.is_default and payload.fieldKey != field.field_key:
        raise fail(status.HTTP_400_BAD_REQUEST, "系统默认字段标识不能修改")
    duplicate = (
        db.query(AdminFeedbackField)
        .filter(
            AdminFeedbackField.form_id == DEFAULT_FEEDBACK_CONFIG_ID,
            AdminFeedbackField.field_key == payload.fieldKey,
            AdminFeedbackField.field_id != fieldId,
        )
        .one_or_none()
    )
    if duplicate is not None:
        raise fail(status.HTTP_400_BAD_REQUEST, "反馈字段标识已存在")
    field.field_key = payload.fieldKey
    field.label = payload.label
    field.type = payload.type
    field.placeholder = payload.placeholder
    field.required = payload.required
    field.options = payload.options
    field.sort = payload.sort
    field.status = payload.status
    field.remark = payload.remark
    field.last_time = utc_now()
    db.commit()
    db.refresh(field)
    return ok(feedback_field_item(field), "反馈字段更新成功")


@router.delete("/feedback/fields/{fieldId}", response_model=AdminResponse, summary="删除反馈字段", description="删除自定义反馈字段；默认字段不能删除。")
def delete_admin_feedback_field(
    fieldId: Annotated[str, Path(description="反馈字段 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    seed_feedback_form_if_empty(db)
    field = db.query(AdminFeedbackField).filter(AdminFeedbackField.field_id == fieldId).one_or_none()
    if field is None:
        raise fail(status.HTTP_404_NOT_FOUND, "反馈字段未找到")
    if field.is_default:
        raise fail(status.HTTP_400_BAD_REQUEST, "系统默认反馈字段不能删除")
    db.delete(field)
    db.commit()
    return ok(None, "反馈字段删除成功")


@router.get("/feedback/submissions", response_model=AdminResponse, summary="反馈记录列表", description="系统配置 - 用户端反馈提交记录列表。")
def list_admin_feedback_submissions(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="搜索手机号、标题、内容、用户 ID")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="处理状态：pending/processing/resolved/rejected")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    query = db.query(FeedbackSubmission)
    if keyword:
        like_keyword = f"%{keyword}%"
        query = query.filter(
            (FeedbackSubmission.phone.ilike(like_keyword))
            | (FeedbackSubmission.title.ilike(like_keyword))
            | (FeedbackSubmission.content.ilike(like_keyword))
            | (FeedbackSubmission.user_id.ilike(like_keyword))
        )
    if status_filter:
        query = query.filter(FeedbackSubmission.status == status_filter)
    total = query.count()
    submissions = query.order_by(FeedbackSubmission.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [feedback_submission_item(item) for item in submissions], "total": total})


@router.get("/feedback/submissions/{feedbackId}", response_model=AdminResponse, summary="反馈记录详情", description="根据反馈 ID 查询用户提交详情。")
def get_admin_feedback_submission(
    feedbackId: Annotated[str, Path(description="反馈 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    feedback = db.query(FeedbackSubmission).filter(FeedbackSubmission.feedback_id == feedbackId).one_or_none()
    if feedback is None:
        raise fail(status.HTTP_404_NOT_FOUND, "反馈记录未找到")
    return ok(feedback_submission_item(feedback))


@router.put("/feedback/submissions/{feedbackId}/status", response_model=AdminResponse, summary="处理反馈记录", description="更新反馈记录处理状态和回复。")
def update_admin_feedback_submission_status(
    feedbackId: Annotated[str, Path(description="反馈 ID")],
    payload: AdminFeedbackStatusPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    feedback = db.query(FeedbackSubmission).filter(FeedbackSubmission.feedback_id == feedbackId).one_or_none()
    if feedback is None:
        raise fail(status.HTTP_404_NOT_FOUND, "反馈记录未找到")
    feedback.status = payload.status
    feedback.reply = payload.reply
    feedback.remark = payload.remark
    feedback.last_time = utc_now()
    db.commit()
    db.refresh(feedback)
    return ok(feedback_submission_item(feedback), "反馈记录已更新")


@router.delete("/feedback/submissions/{feedbackId}", response_model=AdminResponse, summary="删除反馈记录", description="删除用户反馈记录。")
def delete_admin_feedback_submission(
    feedbackId: Annotated[str, Path(description="反馈 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    feedback = db.query(FeedbackSubmission).filter(FeedbackSubmission.feedback_id == feedbackId).one_or_none()
    if feedback is None:
        raise fail(status.HTTP_404_NOT_FOUND, "反馈记录未找到")
    db.delete(feedback)
    db.commit()
    return ok(None, "反馈记录删除成功")


@router.get(
    "/tags",
    response_model=AdminResponse,
    summary="标签列表",
    description="系统配置 - 标签管理列表，支持按标签名称和状态筛选。",
)
def list_admin_tags(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="标签名称关键词")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled 启用，disabled 禁用")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    query = db.query(AdminTag)
    if keyword:
        query = query.filter(AdminTag.name.ilike(f"%{keyword}%"))
    if status_filter:
        query = query.filter(AdminTag.status == status_filter)
    total = query.count()
    tags = query.order_by(AdminTag.sort.asc(), AdminTag.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [tag_item(db, tag) for tag in tags], "total": total})


@router.post(
    "/tags",
    response_model=AdminResponse,
    summary="新增标签",
    description="系统配置 - 新增标签。",
)
def create_admin_tag(
    payload: AdminTagPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    if db.query(AdminTag).filter(AdminTag.name == payload.name).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "标签名称已存在")
    tag = AdminTag(
        tag_id=new_business_id("tag"),
        name=payload.name,
        color=payload.color,
        sort=payload.sort,
        status=payload.status,
        remark=payload.remark,
    )
    db.add(tag)
    db.commit()
    db.refresh(tag)
    return ok(tag_item(db, tag), "标签创建成功")


@router.get(
    "/tags/{tagId}",
    response_model=AdminResponse,
    summary="标签详情",
    description="系统配置 - 查看标签详情。",
)
def get_admin_tag(
    tagId: Annotated[str, Path(description="业务标签 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    tag = db.query(AdminTag).filter(AdminTag.tag_id == tagId).one_or_none()
    if tag is None:
        raise fail(status.HTTP_404_NOT_FOUND, "标签未找到")
    return ok(tag_item(db, tag))


@router.put(
    "/tags/{tagId}",
    response_model=AdminResponse,
    summary="修改标签",
    description="系统配置 - 修改标签信息。",
)
def update_admin_tag(
    tagId: Annotated[str, Path(description="业务标签 ID")],
    payload: AdminTagPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    tag = db.query(AdminTag).filter(AdminTag.tag_id == tagId).one_or_none()
    if tag is None:
        raise fail(status.HTTP_404_NOT_FOUND, "标签未找到")
    duplicate = db.query(AdminTag).filter(AdminTag.name == payload.name, AdminTag.tag_id != tagId).one_or_none()
    if duplicate is not None:
        raise fail(status.HTTP_400_BAD_REQUEST, "标签名称已存在")
    tag.name = payload.name
    tag.color = payload.color
    tag.sort = payload.sort
    tag.status = payload.status
    tag.remark = payload.remark
    tag.last_time = utc_now()
    db.commit()
    db.refresh(tag)
    return ok(tag_item(db, tag), "标签更新成功")


@router.delete(
    "/tags/{tagId}",
    response_model=AdminResponse,
    summary="删除标签",
    description="系统配置 - 删除标签。",
)
def delete_admin_tag(
    tagId: Annotated[str, Path(description="业务标签 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    tag = db.query(AdminTag).filter(AdminTag.tag_id == tagId).one_or_none()
    if tag is None:
        raise fail(status.HTTP_404_NOT_FOUND, "标签未找到")
    db.delete(tag)
    db.commit()
    return ok(None, "标签删除成功")


@router.get(
    "/regions",
    response_model=AdminResponse,
    summary="地区列表",
    description="系统配置 - 地区管理列表，支持按名称、编码、上级地区和状态筛选。",
)
def list_admin_regions(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="地区名称或编码关键词")] = None,
    parentId: Annotated[str | None, Query(description="上级地区 ID")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled 启用，disabled 禁用")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=200, description="每页数量")] = 50,
) -> dict[str, Any]:
    query = db.query(AdminRegion)
    if keyword:
        query = query.filter((AdminRegion.name.ilike(f"%{keyword}%")) | (AdminRegion.code.ilike(f"%{keyword}%")))
    if parentId:
        query = query.filter(AdminRegion.parent_id == parentId)
    if status_filter:
        query = query.filter(AdminRegion.status == status_filter)
    total = query.count()
    regions = query.order_by(AdminRegion.level.asc(), AdminRegion.sort.asc(), AdminRegion.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [region_item(region) for region in regions], "total": total})


@router.post(
    "/regions",
    response_model=AdminResponse,
    summary="新增地区",
    description="系统配置 - 新增地区。",
)
def create_admin_region(
    payload: AdminRegionPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    if db.query(AdminRegion).filter(AdminRegion.code == payload.code).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "地区编码已存在")
    region = AdminRegion(
        region_id=new_business_id("reg"),
        parent_id=payload.parentId,
        name=payload.name,
        code=payload.code,
        level=payload.level,
        sort=payload.sort,
        status=payload.status,
    )
    db.add(region)
    db.commit()
    db.refresh(region)
    return ok(region_item(region), "地区创建成功")


@router.get(
    "/regions/{regionId}",
    response_model=AdminResponse,
    summary="地区详情",
    description="系统配置 - 查看地区详情。",
)
def get_admin_region(
    regionId: Annotated[str, Path(description="业务地区 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    region = db.query(AdminRegion).filter(AdminRegion.region_id == regionId).one_or_none()
    if region is None:
        raise fail(status.HTTP_404_NOT_FOUND, "地区未找到")
    return ok(region_item(region))


@router.put(
    "/regions/{regionId}",
    response_model=AdminResponse,
    summary="修改地区",
    description="系统配置 - 修改地区信息。",
)
def update_admin_region(
    regionId: Annotated[str, Path(description="业务地区 ID")],
    payload: AdminRegionPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    region = db.query(AdminRegion).filter(AdminRegion.region_id == regionId).one_or_none()
    if region is None:
        raise fail(status.HTTP_404_NOT_FOUND, "地区未找到")
    duplicate = db.query(AdminRegion).filter(AdminRegion.code == payload.code, AdminRegion.region_id != regionId).one_or_none()
    if duplicate is not None:
        raise fail(status.HTTP_400_BAD_REQUEST, "地区编码已存在")
    region.parent_id = payload.parentId
    region.name = payload.name
    region.code = payload.code
    region.level = payload.level
    region.sort = payload.sort
    region.status = payload.status
    region.last_time = utc_now()
    db.commit()
    db.refresh(region)
    return ok(region_item(region), "地区更新成功")


@router.delete(
    "/regions/{regionId}",
    response_model=AdminResponse,
    summary="删除地区",
    description="系统配置 - 删除地区；存在下级地区时不允许删除。",
)
def delete_admin_region(
    regionId: Annotated[str, Path(description="业务地区 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    region = db.query(AdminRegion).filter(AdminRegion.region_id == regionId).one_or_none()
    if region is None:
        raise fail(status.HTTP_404_NOT_FOUND, "地区未找到")
    if db.query(AdminRegion).filter(AdminRegion.parent_id == regionId).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "存在下级地区，不能删除")
    db.delete(region)
    db.commit()
    return ok(None, "地区删除成功")


@router.get(
    "/dictionaries",
    response_model=AdminResponse,
    summary="字典列表",
    description="系统配置 - 字典管理列表，支持按字典类型、标签和值筛选；默认返回带 children 的完整树。",
)
def list_admin_dictionaries(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    type: Annotated[str | None, Query(description="字典类型，精确匹配")] = None,
    keyword: Annotated[str | None, Query(description="字典标签或值关键词")] = None,
    parentId: Annotated[str | None, Query(description="上级字典 ID；传入后只查询该字典的直接下级")] = None,
    tree: Annotated[bool, Query(description="是否按父子层级返回 children")] = True,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：enabled 启用，disabled 禁用")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=1000, description="每页数量")] = 1000,
) -> dict[str, Any]:
    query = db.query(AdminDictionary)
    if type:
        query = query.filter(AdminDictionary.type == type)
    if keyword:
        query = query.filter((AdminDictionary.label.ilike(f"%{keyword}%")) | (AdminDictionary.value.ilike(f"%{keyword}%")))
    if parentId is not None:
        query = query.filter(AdminDictionary.parent_id == parentId)
    if status_filter:
        query = query.filter(AdminDictionary.status == status_filter)
    total = query.count()
    dictionaries = query.order_by(AdminDictionary.type.asc(), AdminDictionary.sort.asc(), AdminDictionary.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    if tree and parentId is None:
        return ok({"list": dictionary_tree_items(dictionaries), "total": total})
    return ok({"list": [dictionary_item(dictionary) for dictionary in dictionaries], "total": total})


@router.post(
    "/dictionaries",
    response_model=AdminResponse,
    summary="新增字典",
    description="系统配置 - 新增字典项。",
)
def create_admin_dictionary(
    payload: AdminDictionaryPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    validate_dictionary_parent(db, payload.parentId, payload.type)
    validate_dictionary_unique(db, payload.type, payload.parentId, payload.label, payload.value)
    dictionary = AdminDictionary(
        dict_id=new_business_id("dict"),
        parent_id=payload.parentId,
        type=payload.type,
        label=payload.label,
        value=payload.value,
        sort=payload.sort,
        status=payload.status,
        remark=payload.remark,
    )
    db.add(dictionary)
    db.commit()
    db.refresh(dictionary)
    return ok(dictionary_item(dictionary), "字典创建成功")


@router.get(
    "/dictionaries/{dictId}",
    response_model=AdminResponse,
    summary="字典详情",
    description="系统配置 - 查看字典项详情。",
)
def get_admin_dictionary(
    dictId: Annotated[str, Path(description="业务字典 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    dictionary = get_dictionary_or_404(db, dictId)
    return ok(dictionary_item(dictionary))


@router.get(
    "/dictionaries/{dictId}/children",
    response_model=AdminResponse,
    summary="字典下级列表",
    description="系统配置 - 点击加下级时传入当前字典 ID，查询该字典详情和直接下级字典项。",
)
def list_admin_dictionary_children(
    dictId: Annotated[str, Path(description="业务字典 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    parent = get_dictionary_or_404(db, dictId)
    children = (
        db.query(AdminDictionary)
        .filter(AdminDictionary.parent_id == dictId)
        .order_by(AdminDictionary.sort.asc(), AdminDictionary.create_time.desc())
        .all()
    )
    return ok({"parent": dictionary_item(parent), "list": [dictionary_item(child) for child in children], "total": len(children)})


@router.put(
    "/dictionaries/{dictId}",
    response_model=AdminResponse,
    summary="修改字典",
    description="系统配置 - 修改字典项。",
)
def update_admin_dictionary(
    dictId: Annotated[str, Path(description="业务字典 ID")],
    payload: AdminDictionaryPayload,
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    dictionary = get_dictionary_or_404(db, dictId)
    validate_dictionary_parent(db, payload.parentId, payload.type, self_id=dictId)
    validate_dictionary_unique(db, payload.type, payload.parentId, payload.label, payload.value, self_id=dictId)
    dictionary.parent_id = payload.parentId
    dictionary.type = payload.type
    dictionary.label = payload.label
    dictionary.value = payload.value
    dictionary.sort = payload.sort
    dictionary.status = payload.status
    dictionary.remark = payload.remark
    dictionary.last_time = utc_now()
    db.commit()
    db.refresh(dictionary)
    return ok(dictionary_item(dictionary), "字典更新成功")


@router.delete(
    "/dictionaries/{dictId}",
    response_model=AdminResponse,
    summary="删除字典",
    description="系统配置 - 删除字典项。",
)
def delete_admin_dictionary(
    dictId: Annotated[str, Path(description="业务字典 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    dictionary = get_dictionary_or_404(db, dictId)
    if db.query(AdminDictionary).filter(AdminDictionary.parent_id == dictId).one_or_none():
        raise fail(status.HTTP_400_BAD_REQUEST, "存在下级字典，不能删除")
    db.delete(dictionary)
    db.commit()
    return ok(None, "字典删除成功")


@router.get(
    "/system-messages",
    response_model=AdminResponse,
    summary="系统消息列表",
    description="系统配置 - 系统消息列表，支持按标题、类型、发送范围和状态筛选。",
)
def list_admin_system_messages(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
    keyword: Annotated[str | None, Query(description="消息标题或内容关键词")] = None,
    type: Annotated[str | None, Query(description="消息类型")] = None,
    target: Annotated[str | None, Query(description="发送范围：all/admin/user")] = None,
    status_filter: Annotated[str | None, Query(alias="status", description="状态：draft/published/disabled")] = None,
    page: Annotated[int, Query(ge=1, description="页码")] = 1,
    limit: Annotated[int, Query(ge=1, le=100, description="每页数量")] = 10,
) -> dict[str, Any]:
    query = db.query(AdminSystemMessage)
    if keyword:
        query = query.filter((AdminSystemMessage.title.ilike(f"%{keyword}%")) | (AdminSystemMessage.content.ilike(f"%{keyword}%")))
    if type:
        query = query.filter(AdminSystemMessage.type == type)
    if target:
        query = query.filter(AdminSystemMessage.target == target)
    if status_filter:
        query = query.filter(AdminSystemMessage.status == status_filter)
    total = query.count()
    messages = query.order_by(AdminSystemMessage.is_pinned.desc(), AdminSystemMessage.create_time.desc()).offset((page - 1) * limit).limit(limit).all()
    return ok({"list": [system_message_item(message) for message in messages], "total": total})


@router.post(
    "/system-messages",
    response_model=AdminResponse,
    summary="新增系统消息",
    description="系统配置 - 新增系统消息。",
)
def create_admin_system_message(
    payload: AdminSystemMessagePayload,
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    message = AdminSystemMessage(
        message_id=new_business_id("msg"),
        title=payload.title,
        content=payload.content,
        type=payload.type,
        target=payload.target,
        status=payload.status,
        is_pinned=payload.isPinned,
    )
    db.add(message)
    db.flush()
    ensure_admin_notification(db, admin_subject, message)
    pushed_count = 0
    if message.status == "published":
        pushed_count = publish_system_message_to_users(db, message)
    db.commit()
    db.refresh(message)
    data = system_message_item(message)
    data["pushedCount"] = pushed_count
    return ok(data, "系统消息创建成功")


@router.get(
    "/system-messages/{messageId}",
    response_model=AdminResponse,
    summary="系统消息详情",
    description="系统配置 - 查看系统消息详情。",
)
def get_admin_system_message(
    messageId: Annotated[str, Path(description="业务系统消息 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    message = db.query(AdminSystemMessage).filter(AdminSystemMessage.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "系统消息未找到")
    return ok(system_message_item(message))


@router.put(
    "/system-messages/{messageId}",
    response_model=AdminResponse,
    summary="修改系统消息",
    description="系统配置 - 修改系统消息。",
)
def update_admin_system_message(
    messageId: Annotated[str, Path(description="业务系统消息 ID")],
    payload: AdminSystemMessagePayload,
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    message = db.query(AdminSystemMessage).filter(AdminSystemMessage.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "系统消息未找到")
    old_status = message.status
    message.title = payload.title
    message.content = payload.content
    message.type = payload.type
    message.target = payload.target
    message.status = payload.status
    message.is_pinned = payload.isPinned
    message.last_time = utc_now()
    ensure_admin_notification(db, admin_subject, message)
    pushed_count = 0
    if message.status == "published" and old_status != "published":
        pushed_count = publish_system_message_to_users(db, message)
    db.commit()
    db.refresh(message)
    data = system_message_item(message)
    data["pushedCount"] = pushed_count
    return ok(data, "系统消息更新成功")


@router.put(
    "/system-messages/{messageId}/push",
    response_model=AdminResponse,
    summary="推送系统消息",
    description="系统配置 - 将草稿系统消息发布，并同步生成用户端通知消息。",
)
def push_admin_system_message(
    messageId: Annotated[str, Path(description="业务系统消息 ID")],
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    message = db.query(AdminSystemMessage).filter(AdminSystemMessage.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "系统消息未找到")
    message.status = "published"
    message.last_time = utc_now()
    ensure_admin_notification(db, admin_subject, message)
    pushed_count = publish_system_message_to_users(db, message)
    db.commit()
    db.refresh(message)
    data = system_message_item(message)
    data["pushedCount"] = pushed_count
    return ok(data, "系统消息已推送")


@router.put(
    "/system-messages/{messageId}/retract",
    response_model=AdminResponse,
    summary="撤回系统消息",
    description="系统配置 - 将已发布系统消息撤回为草稿状态。",
)
def retract_admin_system_message(
    messageId: Annotated[str, Path(description="业务系统消息 ID")],
    db: Annotated[Session, Depends(get_db)],
    admin_subject: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    message = db.query(AdminSystemMessage).filter(AdminSystemMessage.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "系统消息未找到")
    message.status = "draft"
    message.last_time = utc_now()
    ensure_admin_notification(db, admin_subject, message)
    db.commit()
    db.refresh(message)
    return ok(system_message_item(message), "系统消息已撤回")


@router.delete(
    "/system-messages/{messageId}",
    response_model=AdminResponse,
    summary="删除系统消息",
    description="系统配置 - 删除系统消息。",
)
def delete_admin_system_message(
    messageId: Annotated[str, Path(description="业务系统消息 ID")],
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[str, Depends(get_admin_subject)],
) -> dict[str, Any]:
    message = db.query(AdminSystemMessage).filter(AdminSystemMessage.message_id == messageId).one_or_none()
    if message is None:
        raise fail(status.HTTP_404_NOT_FOUND, "系统消息未找到")
    db.delete(message)
    db.commit()
    return ok(None, "系统消息删除成功")
